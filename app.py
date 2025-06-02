import os
import pickle
import shutil
import logging
import pytz
import openpyxl  # Already imported to resolve previous NameError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime, timedelta, time, date
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
import face_recognition
import cv2
import numpy as np
from PIL import Image
import io
import base64
import csv
import json
from dotenv import load_dotenv
from pytz import timezone
from calendar import monthrange
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type ,RetryError
from googleapiclient.errors import HttpError
from openpyxl.utils import get_column_letter
import random
import calendar

load_dotenv()


credential_json = os.getenv("GOOGLE_CREDENTIALS_JSON")

if not credential_json:
    raise ValueError("GOOGLE_CREDENTIALS environment variable is not set or empty")


app = Flask(__name__)
app.secret_key = 'your-secret-key'

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

credential_dict=json.loads(credential_json)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
CREDS = Credentials.from_service_account_info(credential_dict, scopes=SCOPES)
service = build('sheets', 'v4', credentials=CREDS)

SHEET_ID = os.getenv('GOOGLE_SHEET_ID')
SUPER_ADMIN_USERNAME = 'ams-hod'
SUPER_ADMIN_PASSWORD = 'admin123'

IMAGES_DIR = 'images'
TEMP_CHECKIN_IMAGES_DIR = 'temp_checkin_images'


# Initialize directories
if not os.path.exists(IMAGES_DIR):
    os.makedirs(IMAGES_DIR)
if not os.path.exists(TEMP_CHECKIN_IMAGES_DIR):
    os.makedirs(TEMP_CHECKIN_IMAGES_DIR)

# @retry(
#     stop=stop_after_attempt(3),
#     wait=wait_exponential(multiplier=1, min=4, max=10),
#     retry=retry_if_exception_type(HttpError),
#     before_sleep=lambda retry_state: logger.info(f"Retrying API call (attempt {retry_state.attempt_number}) due to error: {retry_state.outcome.exception()}"))
# def safe_api_call(api_call):
#     import time
#     time.sleep(1)  # Add a 1-second delay to avoid rate limiting
#     return api_call.execute()

# # def init_sheets():
# #     # Check if Companies sheet exists
# #     try:
# #         safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A1'))
# #     except HttpError as e:
# #         if e.resp.status in [500, 503]:
# #             logger.error(f"Failed to check Companies sheet after retries: {e}")
# #             raise Exception("Google Sheets API is unavailable after retries. Please try again later.")
# #         # If the sheet doesn't exist, create it
# #         safe_api_call(
# #             service.spreadsheets().batchUpdate(
# #                 spreadsheetId=SHEET_ID,
# #                 body={'requests': [{'addSheet': {'properties': {'title': 'Companies'}}}]}
# #             )
# #         )
# #         safe_api_call(
# #             service.spreadsheets().values().update(
# #                 spreadsheetId=SHEET_ID,
# #                 range='Companies!A1:D1',
# #                 valueInputOption='RAW',
# #                 body={'values': [['Company ID', 'Company Name', 'Username', 'Password']]}
# #             )
# #         )
    
# #     # Check if Users sheet exists
# #     try:
# #         safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A1'))
# #     except HttpError as e:
# #         if e.resp.status in [500, 503]:
# #             logger.error(f"Failed to check Users sheet after retries: {e}")
# #             raise Exception("Google Sheets API is unavailable after retries. Please try again later.")
# #         # If the sheet doesn't exist, create it
# #         safe_api_call(
# #             service.spreadsheets().batchUpdate(
# #                 spreadsheetId=SHEET_ID,
# #                 body={'requests': [{'addSheet': {'properties': {'title': 'Users'}}}]}
# #             )
# #         )
# #         safe_api_call(
# #             service.spreadsheets().values().update(
# #                 spreadsheetId=SHEET_ID,
# #                 range='Users!A1:B1',
# #                 valueInputOption='RAW',
# #                 body={'values': [['Company ID', 'User Name']]}
# #             )
# #         )


# def init_sheets():
#     # Check if Companies sheet exists
#     try:
#         safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A1'))
#         logger.info("Companies sheet already exists")
#     except HttpError as e:
#         if e.resp.status == 400 and "Unable to parse range" in str(e):
#             logger.info("Companies sheet does not exist, creating it...")
#             # Create the Companies sheet
#             safe_api_call(
#                 service.spreadsheets().batchUpdate(
#                     spreadsheetId=SHEET_ID,
#                     body={'requests': [{'addSheet': {'properties': {'title': 'Companies'}}}]}
#                 )
#             )
#             # Add headers to the Companies sheet
#             safe_api_call(
#                 service.spreadsheets().values().update(
#                     spreadsheetId=SHEET_ID,
#                     range='Companies!A1:D1',
#                     valueInputOption='RAW',
#                     body={'values': [['Company ID', 'Company Name', 'Username', 'Password']]}
#                 )
#             )
#             logger.info("Companies sheet created successfully")
#         elif e.resp.status in [500, 503]:
#             logger.error(f"Failed to check Companies sheet after retries: {e}")
#             raise Exception("Google Sheets API is unavailable after retries. Please try again later.")
#         else:
#             logger.error(f"Unexpected error when checking Companies sheet: {e}")
#             raise e
    
#     # Check if Users sheet exists
#     try:
#         safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A1'))
#         logger.info("Users sheet already exists")
#     except HttpError as e:
#         if e.resp.status == 400 and "Unable to parse range" in str(e):
#             logger.info("Users sheet does not exist, creating it...")
#             # Create the Users sheet
#             safe_api_call(
#                 service.spreadsheets().batchUpdate(
#                     spreadsheetId=SHEET_ID,
#                     body={'requests': [{'addSheet': {'properties': {'title': 'Users'}}}]}
#                 )
#             )
#             # Add headers to the Users sheet
#             safe_api_call(
#                 service.spreadsheets().values().update(
#                     spreadsheetId=SHEET_ID,
#                     range='Users!A1:B1',
#                     valueInputOption='RAW',
#                     body={'values': [['Company ID', 'User Name']]}
#                 )
#             )
#             logger.info("Users sheet created successfully")
#         elif e.resp.status in [500, 503]:
#             logger.error(f"Failed to check Users sheet after retries: {e}")
#             raise Exception("Google Sheets API is unavailable after retries. Please try again later.")
#         else:
#             logger.error(f"Unexpected error when checking Users sheet: {e}")
#             raise e
        

        # Updated safe_api_call function - only retry server errors (500, 503), not client errors (400)
@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    retry=retry_if_exception_type(lambda e: isinstance(e, HttpError) and e.resp.status in [500, 503]),
    before_sleep=lambda retry_state: logger.info(f"Retrying API call (attempt {retry_state.attempt_number}) due to server error: {retry_state.outcome.exception()}"))
def safe_api_call(api_call):
    import time
    time.sleep(1)  # Add a 1-second delay to avoid rate limiting
    return api_call.execute()

# Alternative approach - create a separate function for sheet checking that doesn't use retry
def check_sheet_exists(sheet_name):
    """Check if a sheet exists without retrying on 400 errors"""
    try:
        service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{sheet_name}!A1').execute()
        return True
    except HttpError as e:
        if e.resp.status == 400 and "Unable to parse range" in str(e):
            return False
        else:
            # For other errors, still raise them
            raise e

# def init_sheets():
#     # Check if Companies sheet exists
#     if not check_sheet_exists('Companies'):
#         logger.info("Companies sheet does not exist, creating it...")
#         try:
#             # Create the Companies sheet
#             safe_api_call(
#                 service.spreadsheets().batchUpdate(
#                     spreadsheetId=SHEET_ID,
#                     body={'requests': [{'addSheet': {'properties': {'title': 'Companies'}}}]}
#                 )
#             )
#             # Add headers to the Companies sheet
#             safe_api_call(
#                 service.spreadsheets().values().update(
#                     spreadsheetId=SHEET_ID,
#                     range='Companies!A1:D1',
#                     valueInputOption='RAW',
#                     body={'values': [['Company ID', 'Company Name', 'Username', 'Password']]}
#                 )
#             )
#             logger.info("Companies sheet created successfully")
#         except Exception as e:
#             logger.error(f"Failed to create Companies sheet: {e}")
#             raise e
#     else:
#         logger.info("Companies sheet already exists")
    
#     # Check if Users sheet exists
#     if not check_sheet_exists('Users'):
#         logger.info("Users sheet does not exist, creating it...")
#         try:
#             # Create the Users sheet
#             safe_api_call(
#                 service.spreadsheets().batchUpdate(
#                     spreadsheetId=SHEET_ID,
#                     body={'requests': [{'addSheet': {'properties': {'title': 'Users'}}}]}
#                 )
#             )
#             # Add headers to the Users sheet
#             safe_api_call(
#                 service.spreadsheets().values().update(
#                     spreadsheetId=SHEET_ID,
#                     range='Users!A1:B1',
#                     valueInputOption='RAW',
#                     body={'values': [['Company ID', 'User Name']]}
#                 )
#             )
#             logger.info("Users sheet created successfully")
#         except Exception as e:
#             logger.error(f"Failed to create Users sheet: {e}")
#             raise e
#     else:
#         logger.info("Users sheet already exists")

def init_sheets():
    # Check if Companies sheet exists
    if not check_sheet_exists('Companies'):
        logger.info("Companies sheet does not exist, creating it...")
        try:
            # Create the Companies sheet
            safe_api_call(
                service.spreadsheets().batchUpdate(
                    spreadsheetId=SHEET_ID,
                    body={'requests': [{'addSheet': {'properties': {'title': 'Companies'}}}]}
                )
            )
            # Add headers to the Companies sheet - Updated to include Leaves Per Month column
            safe_api_call(
                service.spreadsheets().values().update(
                    spreadsheetId=SHEET_ID,
                    range='Companies!A1:E1',
                    valueInputOption='RAW',
                    body={'values': [['Company ID', 'Company Name', 'Username', 'Password', 'Leaves Per Month']]}
                )
            )
            logger.info("Companies sheet created successfully")
        except Exception as e:
            logger.error(f"Failed to create Companies sheet: {e}")
            raise e
    else:
        logger.info("Companies sheet already exists")
    
    # Check if Users sheet exists
    if not check_sheet_exists('Users'):
        logger.info("Users sheet does not exist, creating it...")
        try:
            # Create the Users sheet
            safe_api_call(
                service.spreadsheets().batchUpdate(
                    spreadsheetId=SHEET_ID,
                    body={'requests': [{'addSheet': {'properties': {'title': 'Users'}}}]}
                )
            )
            # Add headers to the Users sheet
            safe_api_call(
                service.spreadsheets().values().update(
                    spreadsheetId=SHEET_ID,
                    range='Users!A1:B1',
                    valueInputOption='RAW',
                    body={'values': [['Company ID', 'User Name']]}
                )
            )
            logger.info("Users sheet created successfully")
        except Exception as e:
            logger.error(f"Failed to create Users sheet: {e}")
            raise e
    else:
        logger.info("Users sheet already exists")
    
    # Check if Leaves sheet exists
    if not check_sheet_exists('Leaves'):
        logger.info("Leaves sheet does not exist, creating it...")
        try:
            # Create the Leaves sheet
            safe_api_call(
                service.spreadsheets().batchUpdate(
                    spreadsheetId=SHEET_ID,
                    body={'requests': [{'addSheet': {'properties': {'title': 'Leaves'}}}]}
                )
            )
            # Add headers to the Leaves sheet
            safe_api_call(
                service.spreadsheets().values().update(
                    spreadsheetId=SHEET_ID,
                    range='Leaves!A1:F1',
                    valueInputOption='RAW',
                    body={'values': [['Company ID', 'User Name', 'Year-Month', 'Leaves Allowed', 'Leaves Taken', 'Leaves Carried']]}
                )
            )
            logger.info("Leaves sheet created successfully")
        except Exception as e:
            logger.error(f"Failed to create Leaves sheet: {e}")
            raise e
    else:
        logger.info("Leaves sheet already exists")


init_sheets()

def generate_company_id(company_name):
    result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:A'))
    existing_ids = result.get('values', [])
    filtered = [cid[0] for cid in existing_ids if cid and cid[0].startswith(company_name + "_")]
    if not filtered:
        return f"{company_name}_01"
    suffixes = [int(cid.split("_")[1]) for cid in filtered if len(cid.split("_")) == 2 and cid.split("_")[1].isdigit()]
    max_suffix = max(suffixes) if suffixes else 0
    return f"{company_name}_{str(max_suffix + 1).zfill(2)}"

def get_sheet_id(sheet_title):
    spreadsheet = safe_api_call(service.spreadsheets().get(spreadsheetId=SHEET_ID))
    sheets = spreadsheet.get('sheets', [])
    for sheet in sheets:
        if sheet.get('properties', {}).get('title') == sheet_title:
            return sheet.get('properties', {}).get('sheetId')
    return None

def load_encodings(company_id):
    encodings_file = f'face_encodings_{company_id}.pkl'
    if os.path.exists(encodings_file):
        with open(encodings_file, 'rb') as f:
            encodings = pickle.load(f)
        logger.debug(f"Loaded encodings for {company_id}: {len(encodings)} users")
        return encodings
    return {}

def save_encodings(company_id, data):
    encodings_file = f'face_encodings_{company_id}.pkl'
    with open(encodings_file, 'wb') as f:
        pickle.dump(data, f)
    logger.debug(f"Saved encodings for {company_id}: {len(data)} users")

def read_attendance_from_sheet(company_id):
    try:
        sheet_data = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A1:Z'))
        sheet_values = sheet_data.get('values', [])
        if not sheet_values or sheet_values[0] == ['Name']:
            return []
        headers = sheet_values[0]
        attendance = []
        for row in sheet_values[1:]:
            name = row[0]
            for i in range(1, len(headers), 4):
                if i + 3 >= len(headers):
                    break
                date = headers[i].replace(' Attendance', '')
                time_range = row[i] if i < len(row) else ''
                expected_checkout = row[i + 1] if i + 1 < len(row) else ''
                hours = row[i + 2] if i + 2 < len(row) else ''
                day_status = row[i + 3] if i + 3 < len(row) else ''
                if time_range:
                    in_time, out_time = parse_time_range(time_range)
                    status = 'Present' if in_time else 'Absent'
                    attendance.append([name, date, in_time, out_time, status, expected_checkout, hours, day_status])
        return attendance
    except Exception as e:
        logger.error(f"Error in read_attendance_from_sheet: {e}")
        return []

def parse_time_range(time_range):
    if '-' in time_range:
        in_time, out_time = time_range.split(' - ')
        return in_time, out_time
    return time_range, ''

def calculate_hours(in_time, out_time):
    if not in_time or not out_time:
        return ''
    try:
        in_time_dt = datetime.strptime(in_time, '%H:%M:%S')
        out_time_dt = datetime.strptime(out_time, '%H:%M:%S')
        time_diff = out_time_dt - in_time_dt
        if time_diff.total_seconds() < 0:
            time_diff += timedelta(days=1)
        hours = time_diff.total_seconds() / 3600
        return f"{hours:.2f}"
    except ValueError:
        return ''

def update_sheet(company_id, attendance):
    try:
        today = datetime.now().strftime('%d/%m/%Y')
        headers = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A1:1')).get('values', [[]])[0]
        if not headers or headers == ['Name']:
            headers = ['Name']
        else:
            existing_dates = set()
            for header in headers[1:]:
                if header.endswith(' Attendance'):
                    existing_dates.add(header.replace(' Attendance', ''))
            headers = ['Name']
            for date in sorted(existing_dates):
                headers.extend([f"{date} Attendance", f"{date} Check-out", f"{date} Hours", f"{date} Day Status"])

        existing_dates = {h.replace(' Attendance', '') for h in headers if h.endswith(' Attendance')}
        for date in set(r[1] for r in attendance):
            if date not in existing_dates:
                headers.extend([f"{date} Attendance", f"{date} Check-out", f"{date} Hours", f"{date} Day Status"])

        if f"{today} Attendance" not in headers:
            headers.extend([f"{today} Attendance", f"{today} Check-out", f"{today} Hours", f"{today} Day Status"])

        all_names = list(load_encodings(company_id).keys())
        updated_data = []
        for name in all_names:
            row = [name]
            for i in range(1, len(headers), 4):
                date = headers[i].replace(' Attendance', '')
                time_range = ''
                check_out = ''
                hours = ''
                day_status = ''
                for record in attendance:
                    if record[0] == name and record[1] == date:
                        in_time = record[2] if record[2] else ''
                        out_time = record[3] if record[3] else ''
                        day_status = record[7] if len(record) > 7 else ''
                        if in_time and out_time:
                            time_range = f"{in_time} - {out_time}"
                            hours = calculate_hours(in_time, out_time)
                        elif in_time:
                            time_range = in_time
                        check_out = out_time
                row.extend([time_range, check_out, hours, day_status])
            updated_data.append(row)

        safe_api_call(
            service.spreadsheets().values().update(
                spreadsheetId=SHEET_ID,
                range=f'{company_id}!A1',
                valueInputOption='RAW',
                body={'values': [headers]}
            )
        )
        if updated_data:
            safe_api_call(
                service.spreadsheets().values().update(
                    spreadsheetId=SHEET_ID,
                    range=f'{company_id}!A2',
                    valueInputOption='RAW',
                    body={'values': updated_data}
                )
            )
    except Exception as e:
        logger.error(f"Error in update_sheet: {e}")
        raise

def get_checkin_image_base64(company_id, name, date_str):
    image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{date_str.replace('/', '-')}.jpg")
    if os.path.exists(image_path):
        with open(image_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            return f"data:image/jpeg;base64,{encoded_string}"
    return None

def log_attendance(company_id, name, action):
    now = datetime.now() + india_offset
    date_str = now.strftime('%d/%m/%Y')
    time_str = now.strftime('%H:%M:%S')

    attendance = read_attendance_from_sheet(company_id)
    today_records = [r for r in attendance if r[0] == name and r[1] == date_str]

    if not today_records:
        if action == 'checkin':
            checkin_time = datetime.strptime(time_str, '%H:%M:%S').time() 
            checkin_dt = datetime.combine(datetime.today(), checkin_time) + india_offset
            time_10_00 = datetime.combine(datetime.today(), time(10, 0)) + india_offset
            time_10_30 = datetime.combine(datetime.today(), time(10, 30)) + india_offset
            time_11_00 = datetime.combine(datetime.today(), time(11, 0)) + india_offset
            day_status = 'Full Day'
            expected_checkout = '18:30:00'

            if time_10_00 <= checkin_dt < time_10_30:
                expected_checkout = '18:30:00'
            elif time_10_30 <= checkin_dt <= time_11_00:
                checkout_dt = checkin_dt + timedelta(hours=8)
                expected_checkout = checkout_dt.strftime('%H:%M:%S')
            elif checkin_dt > time_11_00:
                expected_checkout = '18:30:00'
                day_status = 'Half Day'

            attendance.append([name, date_str, time_str, '', 'Present', expected_checkout, '', day_status])
            update_sheet(company_id, attendance)
            return True, expected_checkout, day_status
    else:
        last_record = today_records[-1]
        checkin_time = datetime.strptime(last_record[2], '%H:%M:%S') if last_record[2] else None
        checkout_time = datetime.strptime(last_record[3], '%H:%M:%S') if last_record[3] else None
        day_status = last_record[7] if len(last_record) > 7 else 'Full Day'

        if action == 'checkout' and checkin_time and not checkout_time:
            time_since_checkin = now - datetime.combine(date.today(), checkin_time.time()) + india_offset
            expected_checkout = datetime.strptime(last_record[5], '%H:%M:%S') if last_record[5] else datetime.strptime('18:30:00', '%H:%M:%S')
            if time_since_checkin.total_seconds() < 0:
                time_since_checkin += timedelta(days=1)
            if time_since_checkin >= timedelta(hours=7):
                last_record[3] = time_str
                last_record[4] = 'Present'
                last_record[6] = calculate_hours(last_record[2], time_str)
                update_sheet(company_id, attendance)
                return True, last_record[6], day_status
            return False, last_record[5], day_status
        elif action == 'checkin' and not checkout_time:
            if not checkin_time:
                last_record[2] = time_str
                last_record[4] = 'Present'
                checkin_dt = datetime.strptime(time_str, '%H:%M:%S') + india_offset
                time_10_00 = datetime.combine(datetime.today(), time(10, 0)) + india_offset
                time_10_30 = datetime.combine(datetime.today(), time(10, 30)) + india_offset
                time_11_00 = datetime.combine(datetime.today(), time(11, 0)) + india_offset
                day_status = 'Full Day'
                expected_checkout = '18:30:00'

                if time_10_00 <= checkin_dt < time_10_30:
                    expected_checkout = '18:30:00'
                elif time_10_30 <= checkin_dt <= time_11_00:
                    expected_checkout = (checkin_dt + timedelta(hours=8)).strftime('%H:%M:%S')
                elif checkin_dt > time_11_00:
                    expected_checkout = '18:30:00'
                    day_status = 'Half Day'

                last_record[5] = expected_checkout
                last_record[7] = day_status
                update_sheet(company_id, attendance)
                return True, expected_checkout, day_status
    return False, '', day_status

def find_best_match(face_encoding, known_faces, tolerance=0.5, strict_threshold=0.45):
    matches = []
    for current_name, known_encodings in known_faces.items():
        distances = face_recognition.face_distance(known_encodings, face_encoding)
        for i, distance in enumerate(distances):
            if distance < tolerance:
                matches.append((current_name, distance))
    
    if matches:
        best_match = min(matches, key=lambda x: x[1])
        name, distance = best_match
        if distance < strict_threshold:
            logger.debug(f"Found match: {name} with distance {distance}")
            return name, distance
        else:
            logger.debug(f"Best match {name} rejected: distance {distance} >= {strict_threshold}")
    
    logger.debug("No valid match found for face.")
    return None

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        if username == SUPER_ADMIN_USERNAME and password == SUPER_ADMIN_PASSWORD:
            session['username'] = username
            session['role'] = 'super_admin'
            return redirect(url_for('super_admin'))
        result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D'))
        users = result.get('values', [])
        for user in users:
            if len(user) >= 4 and user[2].strip() == username and user[3].strip() == password:
                session['username'] = username
                session['role'] = 'company_admin'
                session['company_id'] = user[0]
                session['company_name'] = user[1]
                user_result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A2:B'))
                company_users = [row[1] for row in user_result.get('values', []) if row and row[0] == user[0]]
                if not company_users:
                    return redirect(url_for('admin_panel'))
                return redirect(url_for('admin_panel'))
        return render_template('login.html', error='Invalid credentials')
    return render_template('login.html')

@app.route('/super_admin', methods=['GET', 'POST'])
def super_admin():
    if 'username' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        if 'delete_id' in request.form:
            delete_id = request.form['delete_id']
            sheet_data = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:A'))
            values = sheet_data.get('values', [])
            row_index = None
            for i, row in enumerate(values, start=2):
                if row and row[0] == delete_id:
                    row_index = i
                    break
            if row_index:
                requests = [{
                    "deleteDimension": {
                        "range": {
                            "sheetId": get_sheet_id('Companies'),
                            "dimension": "ROWS",
                            "startIndex": row_index - 1,
                            "endIndex": row_index
                        }
                    }
                }]
                safe_api_call(service.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={'requests': requests}))
                user_data = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A2:B'))
                user_values = user_data.get('values', [])
                updated_users = [row for row in user_values if row and row[0] != delete_id]
                safe_api_call(
                    service.spreadsheets().values().update(
                        spreadsheetId=SHEET_ID,
                        range='Users!A2:B',
                        valueInputOption='RAW',
                        body={'values': updated_users}
                    )
                )
                encodings_file = f'face_encodings_{delete_id}.pkl'
                if os.path.exists(encodings_file):
                    os.remove(encodings_file)
                company_dir = os.path.join(IMAGES_DIR, delete_id)
                if os.path.exists(company_dir):
                    shutil.rmtree(company_dir)
        else:
            company_name = request.form['company_name'].strip()
            company_username = request.form['company_username'].strip()
            company_password = request.form['company_password'].strip()
            company_id = generate_company_id(company_name)
            safe_api_call(
                service.spreadsheets().values().append(
                    spreadsheetId=SHEET_ID,
                    range='Companies!A2:D2',
                    valueInputOption='RAW',
                    body={'values': [[company_id, company_name, company_username, company_password]]}
                )
            )

    result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D'))
    companies = result.get('values', [])
    user_result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A2:B'))
    users = {company[0]: [] for company in companies}
    for row in user_result.get('values', []):
        if row and len(row) >= 2:
            users[row[0]].append(row[1])

    return render_template('super_admin.html', companies=companies, users=users)

@app.route('/edit_company/<company_id>', methods=['GET', 'POST'])
def edit_company(company_id):
    if 'username' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))

    # Fetch all companies
    companies_data = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D')).get('values', [])

    # Find the row for the company to edit
    row_index = None
    company_row = None
    for i, row in enumerate(companies_data, start=2):
        if row[0] == company_id:
            row_index = i
            company_row = row
            break

    if request.method == 'POST' and row_index:
        new_name = request.form['company_name'].strip()
        new_username = request.form['company_username'].strip()
        new_password = request.form['company_password'].strip()

        safe_api_call(service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range=f'Companies!B{row_index}:D{row_index}',
            valueInputOption='RAW',
            body={'values': [[new_name, new_username, new_password]]}
        ))

        return redirect(url_for('super_admin'))

    return render_template('edit_company.html', company_id=company_id, company=company_row)


@app.route('/combined_dashboard')
def combined_dashboard():
    if 'username' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))

    # Fetch companies
    result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D'))
    companies_raw = result.get('values', [])
    
    # Fetch users
    user_result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A2:B'))
    users = {company[0]: [] for company in companies_raw}
    for row in user_result.get('values', []):
        if row and len(row) >= 2:
            users[row[0]].append(row[1])

    # Calculate attendance stats
        # Calculate attendance stats
    today = datetime.now().strftime('%d/%m/%Y')
    companies_with_stats = []
    for company in companies_raw:
        company_id = company[0]
        total_users = len(users.get(company_id, []))
        attendance = read_attendance_from_sheet(company_id)
        todays_attendance = [r for r in attendance if r[1] == today]
        present_today = len([r for r in todays_attendance if r[4] == 'Present'])
        absent_today = total_users - present_today
        companies_with_stats.append({
            0: company[0],  # Company ID
            1: company[1],  # Company Name
            'total_users': total_users,
            'present_today': present_today,
            'absent_today': absent_today
        })

    # Create a version of companies_with_stats with string keys for JSON serialization
    company_stats_for_json = [
        {
            '0': company[0],
            '1': company[1],
            'total_users': company['total_users'],
            'present_today': company['present_today'],
            'absent_today': company['absent_today']
        }
        for company in companies_with_stats
    ]

    return render_template('combined_dashboard.html', companies=companies_with_stats, companyStats=company_stats_for_json)

# @app.route('/company_dashboard/<company_id>')
# def company_dashboard(company_id):
#     if 'username' not in session or session.get('role') != 'super_admin':
#         return redirect(url_for('login'))

#     # Fetch company name
#     result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D'))
#     companies = result.get('values', [])
#     company = next((c for c in companies if c[0] == company_id), None)
#     if not company:
#         return "Company not found", 404
#     company_name = company[1]

#     # Fetch attendance records
#     attendance = read_attendance_from_sheet(company_id)
#     attendance_records = [
#         {
#             'name': record[0],
#             'date': record[1],
#             'time_range': f"{record[2]} - {record[3]}" if record[2] and record[3] else (record[2] or record[3] or 'N/A'),
#             'status': record[4],
#             'hours': record[6] or 'N/A',
#             'day_status': record[7] or 'N/A'
#         } for record in attendance
#     ]

#     # Get unique user names and attendance summaries
#     users_list = sorted(list(set(record[0] for record in attendance)))
#     user_attendance_data = {}
#     for user in users_list:
#         user_records = [r for r in attendance if r[0] == user]
#         unique_dates = set(r[1] for r in user_records)  # Unique dates for this user
#         present_days = len([r for r in user_records if r[4] == 'Present'])
#         absent_days = len(unique_dates) - present_days  # Count days with no "Present" status
#         leaves_taken = min(absent_days, 2)  # Cap leaves at 2
#         user_attendance_data[user] = {
#             'present_days': present_days,
#             'absent_days': absent_days,
#             'leaves_taken': leaves_taken
#         }

#     return render_template(
#         'company_dashboard.html',
#         company_id=company_id,
#         company_name=company_name,
#         attendance_records=attendance_records,
#         users_list=users_list,
#         user_attendance_data=user_attendance_data
#     )

# @app.route('/company_dashboard/<company_id>')
# def company_dashboard(company_id):
#     if 'username' not in session or session.get('role') != 'super_admin':
#         return redirect(url_for('login'))

#     # Fetch company name
#     result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D').execute()
#     companies = result.get('values', [])
#     company = next((c for c in companies if c[0] == company_id), None)
#     if not company:
#         return "Company not found", 404
#     company_name = company[1]

#     # Get filter parameters
#     filter_name = request.args.get('filter_name', '')
#     filter_month = request.args.get('filter_month', '')
#     filter_date = request.args.get('filter_date', '')
#     filter_start_date = request.args.get('filter_start_date', '')
#     filter_end_date = request.args.get('filter_end_date', '')

#     # Fetch attendance records
#     attendance = read_attendance_from_sheet(company_id)
    
#     # Get unique user names and available months
#     users_list = sorted(list(set(record[0] for record in attendance)))
#     unique_dates = set(record[1] for record in attendance)
#     available_months = []
#     month_names = {
#         '01': 'January', '02': 'February', '03': 'March', '04': 'April',
#         '05': 'May', '06': 'June', '07': 'July', '08': 'August',
#         '09': 'September', '10': 'October', '11': 'November', '12': 'December'
#     }
#     months_set = set()
#     for date_str in unique_dates:
#         try:
#             month = date_str.split('/')[1]
#             if month in month_names and month not in months_set:
#                 available_months.append({'name': month_names[month], 'value': month})
#                 months_set.add(month)
#         except (IndexError, ValueError):
#             continue
#     available_months.sort(key=lambda x: x['value'])

#     # Filter attendance records
#     filtered_attendance = attendance
#     if filter_name:
#         filtered_attendance = [r for r in filtered_attendance if r[0] == filter_name]
#     if filter_month:
#         filtered_attendance = [r for r in filtered_attendance if r[1].split('/')[1] == filter_month]
#     if filter_date:
#         try:
#             # Convert filter_date (YYYY-MM-DD) to DD/MM/YYYY
#             date_obj = datetime.strptime(filter_date, '%Y-%m-%d')
#             formatted_date = date_obj.strftime('%d/%m/%Y')
#             filtered_attendance = [r for r in filtered_attendance if r[1] == formatted_date]
#         except ValueError:
#             filtered_attendance = []
#     if filter_start_date and filter_end_date:
#         try:
#             start_date = datetime.strptime(filter_start_date, '%Y-%m-%d')
#             end_date = datetime.strptime(filter_end_date, '%Y-%m-%d')
#             if start_date > end_date:
#                 start_date, end_date = end_date, start_date
#             filtered_attendance = [
#                 r for r in filtered_attendance
#                 if start_date <= datetime.strptime(r[1], '%d/%m/%Y') <= end_date
#             ]
#         except ValueError:
#             filtered_attendance = []

#     attendance_records = [
#         {
#             'name': record[0],
#             'date': record[1],
#             'time_range': f"{record[2]} - {record[3]}" if record[2] and record[3] else (record[2] or record[3] or 'N/A'),
#             'status': record[4],
#             'hours': record[6] or 'N/A',
#             'day_status': record[7] or 'N/A'
#         } for record in filtered_attendance
#     ]

#     # Get attendance summaries for chart (unfiltered)
#     user_attendance_data = {}
#     for user in users_list:
#         user_records = [r for r in attendance if r[0] == user]
#         unique_dates = set(r[1] for r in user_records)
#         present_days = len([r for r in user_records if r[4] == 'Present'])
#         absent_days = len(unique_dates) - present_days
#         leaves_taken = min(absent_days, 2)
#         user_attendance_data[user] = {
#             'present_days': present_days,
#             'absent_days': absent_days,
#             'leaves_taken': leaves_taken
#         }

#     return render_template(
#         'company_dashboard.html',
#         company_id=company_id,
#         company_name=company_name,
#         attendance_records=attendance_records,
#         users_list=users_list,
#         user_attendance_data=user_attendance_data,
#         available_months=available_months,
#         filter_name=filter_name,
#         filter_month=filter_month,
#         filter_date=filter_date,
#         filter_start_date=filter_start_date,
#         filter_end_date=filter_end_date
#     )

from datetime import datetime, timedelta
from flask import session, redirect, url_for, request, render_template
from googleapiclient.discovery import build

# Assuming `service` and `SHEET_ID` are defined elsewhere
# Assuming `read_attendance_from_sheet` returns list of [name, date, check_in, check_out, status, ..., hours, day_status]

@app.route('/company_dashboard/<company_id>')
def company_dashboard(company_id):
    if 'username' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))

    # Fetch company name
    result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D').execute()
    companies = result.get('values', [])
    company = next((c for c in companies if c[0] == company_id), None)
    if not company:
        return "Company not found", 404
    company_name = company[1]

    # Get filter parameters
    filter_name = request.args.get('filter_name', '')
    filter_month = request.args.get('filter_month', '')
    filter_date = request.args.get('filter_date', '')

    # Fetch attendance records
    attendance = read_attendance_from_sheet(company_id)
    
    # Get unique user names
    users_list = sorted(list(set(record[0] for record in attendance)))

    # Define expected working days (e.g., last 30 days, excluding weekends)
    def get_working_days(start_date, end_date):
        working_days = []
        current_date = start_date
        while current_date <= end_date:
            # Exclude weekends (Saturday=5, Sunday=6)
            if current_date.weekday() < 5:
                working_days.append(current_date.strftime('%d/%m/%Y'))
            current_date += timedelta(days=1)
        return working_days

    # Set date range for expected working days (e.g., last 30 days)
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    working_days = get_working_days(start_date, end_date)

    # Get unique months for filter dropdown
    unique_dates = set(record[1] for record in attendance)
    available_months = []
    month_names = {
        '01': 'January', '02': 'February', '03': 'March', '04': 'April',
        '05': 'May', '06': 'June', '07': 'July', '08': 'August',
        '09': 'September', '10': 'October', '11': 'November', '12': 'December'
    }
    months_set = set()
    for date_str in unique_dates:
        try:
            month = date_str.split('/')[1]
            if month in month_names and month not in months_set:
                available_months.append({'name': month_names[month], 'value': month})
                months_set.add(month)
        except (IndexError, ValueError):
            continue
    available_months.sort(key=lambda x: x['value'])

    # Create a comprehensive attendance list with absent records
    comprehensive_attendance = []
    for user in users_list:
        user_records = [r for r in attendance if r[0] == user]
        user_dates = set(r[1] for r in user_records)
        
        # Add present records
        comprehensive_attendance.extend(user_records)
        
        # Add absent records for missing working days
        for date in working_days:
            if date not in user_dates:
                # Create an absent record
                comprehensive_attendance.append([
                    user,  # record[0]: name
                    date,  # record[1]: date
                    '',    # record[2]: check_in (empty)
                    '',    # record[3]: check_out (empty)
                    'Absent',  # record[4]: status
                    '',    # record[5]: (assuming this is unused)
                    '0',   # record[6]: hours
                    'Absent'  # record[7]: day_status
                ])

    # Filter attendance records
    filtered_attendance = comprehensive_attendance
    if filter_name:
        filtered_attendance = [r for r in filtered_attendance if r[0] == filter_name]
    if filter_month:
        filtered_attendance = [r for r in filtered_attendance if r[1].split('/')[1] == filter_month]
    if filter_date:
        try:
            date_obj = datetime.strptime(filter_date, '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d/%m/%Y')
            filtered_attendance = [r for r in filtered_attendance if r[1] == formatted_date]
        except ValueError:
            filtered_attendance = []

    # Create attendance records for display
    attendance_records = [
        {
            'name': record[0],
            'date': record[1],
            'time_range': f"{record[2]} - {record[3]}" if record[2] and record[3] else (record[2] or record[3] or 'N/A'),
            'status': record[4],
            'hours': record[6] or 'N/A',
            'day_status': record[7] or 'N/A'
        } for record in filtered_attendance
    ]

    # Calculate attendance summaries for chart
    user_attendance_data = {}
    for user in users_list:
        user_records = [r for r in comprehensive_attendance if r[0] == user]
        present_days = len([r for r in user_records if r[4].lower() == 'present'])
        absent_days = len([r for r in user_records if r[4].lower() == 'absent'])
        leaves_taken = min(absent_days, 2)  # Assuming max 2 leaves allowed

        user_attendance_data[user] = {
            'present_days': present_days,
            'absent_days': absent_days,
            'leaves_taken': leaves_taken
        }

    return render_template(
        'company_dashboard.html',
        company_id=company_id,
        company_name=company_name,
        attendance_records=attendance_records,
        users_list=users_list,
        user_attendance_data=user_attendance_data,
        available_months=available_months,
        filter_name=filter_name,
        filter_month=filter_month,
        filter_date=filter_date
    )

# @app.route('/student_dashboard/<company_id>/<name>')
# def student_dashboard(company_id, name):
#     if 'username' not in session or session.get('role') != 'super_admin':
#         return redirect(url_for('login'))

#     # Fetch company name
#     result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D'))
#     companies = result.get('values', [])
#     company = next((c for c in companies if c[0] == company_id), None)
#     if not company:
#         return "Company not found", 404
#     company_name = company[1]

#     # Fetch attendance records for the user
#     attendance = read_attendance_from_sheet(company_id)
#     attendance_records = [
#         {
#             'date': record[1],
#             'time_range': f"{record[2]} - {record[3]}" if record[2] and record[3] else (record[2] or record[3] or 'N/A'),
#             'status': record[4],
#             'hours': record[6] or 'N/A',
#             'day_status': record[7] or 'N/A'
#         } for record in attendance if record[0] == name
#     ]

#     return render_template('student_dashboard.html', company_id=company_id, company_name=company_name, student_name=name, attendance_records=attendance_records)


# @app.route('/student_dashboard/<company_id>/<name>')
# def student_dashboard(company_id, name):
#     if 'username' not in session or session.get('role') != 'super_admin':
#         return redirect(url_for('login'))

#     # Fetch company name
#     result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D').execute()
#     companies = result.get('values', [])
#     company = next((c for c in companies if c[0] == company_id), None)
#     if not company:
#         return "Company not found", 404
#     company_name = company[1]

#     # Fetch attendance records
#     attendance = read_attendance_from_sheet(company_id)
#     user_records = [r for r in attendance if r[0] == name]
#     attendance_records = [
#         {
#             'date': record[1],
#             'time_range': f"{record[2]} - {record[3]}" if record[2] and record[3] else (record[2] or record[3] or 'N/A'),
#             'status': record[4],
#             'hours': record[6] or 'N/A',
#             'day_status': record[7] or 'N/A'
#         } for record in user_records
#     ]

#     # Calculate attendance for the last month
#     today = datetime.now()
#     current_year_month = today.strftime("%Y-%m")
#     last_month_end = (today.replace(day=1) - timedelta(days=1))
#     last_month_start = last_month_end.replace(day=1)
#     total_days = last_month_end.day  # 28, 29, 30, or 31

#     dates = [(last_month_start + timedelta(days=i)).strftime('%d/%m/%Y') for i in range(total_days)]
#     statuses = []
#     present_days = 0
#     for date_str in dates:
#         record = next((r for r in user_records if r[1] == date_str), None)
#         if record and record[4] == 'Present':
#             statuses.append(1)
#             present_days += 1
#         else:
#             statuses.append(0)

#     attendance_data = {
#         'dates': dates,
#         'statuses': statuses
#     }

#     # Get leave balance for current month
#     leaves_allowed, leaves_taken = get_leave_balance(company_id, name, current_year_month)

#     return render_template(
#         'student_dashboard.html',
#         company_id=company_id,
#         company_name=company_name,
#         student_name=name,
#         attendance_records=attendance_records,
#         attendance_data=attendance_data,
#         present_days=present_days,
#         total_days=total_days,
#         leaves_taken=leaves_taken,
#         leaves_allowed=leaves_allowed
#     )


# @app.route('/student_dashboard/<company_id>/<name>')
# def student_dashboard(company_id, name):
#     if 'username' not in session or session.get('role') != 'super_admin':
#         return redirect(url_for('login'))

#     # Fetch company name
#     result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:E').execute()
#     companies = result.get('values', [])
#     company = next((c for c in companies if c[0] == company_id), None)
#     if not company:
#         return "Company not found", 404
#     company_name = company[1]

#     # Fetch attendance records
#     attendance = read_attendance_from_sheet(company_id)
#     user_records = [r for r in attendance if r[0] == name]
#     attendance_records = [
#         {
#             'date': record[1],
#             'time_range': f"{record[2]} - {record[3]}" if record[2] and record[3] else (record[2] or record[3] or 'N/A'),
#             'status': record[4],
#             'hours': record[6] or 'N/A',
#             'day_status': record[7] or 'N/A'
#         } for record in user_records
#     ]

#     # Calculate attendance for the current month
#     today = datetime.now()
#     current_year_month = today.strftime("%Y-%m")
#     year, month = today.year, today.month
#     _, total_days = calendar.monthrange(year, month)  # 31 for May
#     current_month_start = today.replace(day=1)
#     current_month_end = current_month_start + timedelta(days=total_days - 1)

#     dates = [(current_month_start + timedelta(days=i)).strftime('%d/%m/%Y') for i in range(total_days)]
#     # Initialize data arrays
#     present_data = [0] * total_days
#     absent_data = [0] * total_days
#     halfday_data = [0] * total_days
#     leave_data = [0] * total_days
#     present_days = 0

#     # Fetch leave records
#     leave_result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Leaves!A2:F').execute()
#     leave_records = leave_result.get('values', [])
#     leave_dates = []
#     for record in leave_records:
#         if (len(record) >= 3 and record[0] == company_id and record[1] == name and 
#             record[2] == current_year_month and len(record) > 4 and record[4]):
#             try:
#                 leaves_taken = int(record[4])
#                 for i in range(leaves_taken):
#                     if i < total_days:
#                         leave_dates.append(dates[i])
#             except ValueError:
#                 pass

#     # Process attendance up to current day
#     for i, date_str in enumerate(dates):
#         day_of_month = i + 1
#         if day_of_month > today.day:
#             continue

#         if date_str in leave_dates:
#             leave_data[i] = random.uniform(0.1, 0.9)
#         else:
#             record = next((r for r in user_records if r[1] == date_str), None)
#             if record and record[4] == 'Present':
#                 present_data[i] = random.uniform(0.1, 0.9)
#                 present_days += 1
#             elif record and record[4] == 'HalfDay':
#                 halfday_data[i] = random.uniform(0.1, 0.9)
#             elif record and record[4] == 'Leave':
#                 leave_data[i] = random.uniform(0.1, 0.9)
#             else:
#                 absent_data[i] = random.uniform(0.1, 0.9)

#     attendance_data = {
#         'dates': dates,
#         'present': present_data,
#         'absent': absent_data,
#         'halfday': halfday_data,
#         'leave': leave_data
#     }

#     # Get leave balance
#     leaves_allowed, leaves_taken = get_leave_balance(company_id, name, current_year_month)

#     logger.debug(f"Generated attendance data for {name}: {len(dates)} days, Present: {sum(1 for x in present_data if x > 0)}, Absent: {sum(1 for x in absent_data if x > 0)}, HalfDay: {sum(1 for x in halfday_data if x > 0)}, Leave: {sum(1 for x in leave_data if x > 0)}, Total Days: {total_days}")

#     return render_template(
#         'student_dashboard.html',
#         company_id=company_id,
#         company_name=company_name,
#         student_name=name,
#         attendance_records=attendance_records,
#         attendance_data=attendance_data,
#         present_days=present_days,
#         total_days=total_days,
#         leaves_taken=leaves_taken,
#         leaves_allowed=leaves_allowed,
#         month_name=today.strftime('%B %Y'),
#         current_day=today.day,
#         current_month_year=today.strftime('%m/%Y'),
#         current_day_index=today.day - 1
#     )

@app.route('/student_dashboard/<company_id>/<name>')
def student_dashboard(company_id, name):
    if 'username' not in session or session.get('role') != 'super_admin':
        return redirect(url_for('login'))

    # Fetch company name and leaves per month
    result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:E').execute()
    companies = result.get('values', [])
    company = next((c for c in companies if c[0] == company_id), None)
    if not company:
        return "Company not found", 404
    company_name = company[1]
    leaves_per_month = int(company[4]) if len(company) > 4 and company[4] else 2

    # Get filter parameters (month and year)
    current_date = datetime.now(timezone('Asia/Kolkata'))  # Current date: May 29, 2025
    month = int(request.args.get('month', current_date.month))  # Default to May (5)
    year = int(request.args.get('year', current_date.year))  # Default to 2025

    # Calculate total days in the selected month
    _, total_days = calendar.monthrange(year, month)

    # Define the selected month and year as a datetime object
    selected_month_start = datetime(year, month, 1)
    month_name = selected_month_start.strftime('%B')  # e.g., "May" or "June"
    year_month = selected_month_start.strftime('%Y-%m')  # e.g., "2025-05"

    # Fetch attendance records
    attendance = read_attendance_from_sheet(company_id)
    user_records = [r for r in attendance if r[0] == name]
    
    # Filter records for the selected month and year
    filtered_records = [
        r for r in user_records 
        if datetime.strptime(r[1], '%d/%m/%Y').strftime('%Y-%m') == year_month
    ]

    attendance_records = [
        {
            'date': record[1],
            'time_range': f"{record[2]} - {record[3]}" if record[2] and record[3] else (record[2] or record[3] or 'N/A'),
            'status': record[4],
            'hours': record[6] or 'N/A',
            'day_status': record[7] or 'N/A'
        } for record in filtered_records
    ]

    # Generate dates for the selected month
    dates = [(selected_month_start + timedelta(days=i)).strftime('%d/%m/%Y') for i in range(total_days)]
    
    # Initialize data arrays for the chart
    present_data = [0] * total_days
    absent_data = [0] * total_days
    halfday_data = [0] * total_days
    leave_data = [0] * total_days
    present_days = 0

    # Fetch leave records for the selected month
    leave_result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Leaves!A2:F').execute()
    leave_records = leave_result.get('values', [])
    leave_dates = []
    for record in leave_records:
        if (len(record) >= 3 and record[0] == company_id and record[1] == name and 
            record[2] == year_month and len(record) > 4 and record[4]):
            try:
                leaves_taken = int(record[4])
                for i in range(leaves_taken):
                    if i < total_days:
                        leave_dates.append(dates[i])
            except ValueError:
                pass

    # Determine the last marked day
    last_marked_day = total_days  # Default to the last day of the selected month
    if year == current_date.year and month == current_date.month:
        last_marked_day = current_date.day  # Use current day (29) for May 2025
    elif year > current_date.year or (year == current_date.year and month > current_date.month):
        last_marked_day = 0  # No data for future months

    # Process attendance up to the last marked day
    for i, date_str in enumerate(dates):
        day_of_month = i + 1
        if day_of_month > last_marked_day:
            continue

        if date_str in leave_dates:
            leave_data[i] = random.uniform(0.1, 0.9)
            logger.debug(f"Marked {date_str} as Leave for {name}")
        else:
            record = next((r for r in filtered_records if r[1] == date_str), None)
            if record:
                if record[4] == 'Present':
                    present_data[i] = random.uniform(0.1, 0.9)
                    present_days += 1
                    logger.debug(f"Marked {date_str} as Present for {name}")
                elif record[4] == 'HalfDay':
                    halfday_data[i] = random.uniform(0.1, 0.9)
                    logger.debug(f"Marked {date_str} as HalfDay for {name}")
                elif record[4] == 'Leave':
                    leave_data[i] = random.uniform(0.1, 0.9)
                    logger.debug(f"Marked {date_str} as Leave for {name}")
                else:
                    absent_data[i] = random.uniform(0.1, 0.9)
                    logger.debug(f"Marked {date_str} as Absent for {name} (status: {record[4]})")
            else:
                absent_data[i] = random.uniform(0.1, 0.9)
                logger.debug(f"Marked {date_str} as Absent for {name} (no record)")

    # Prepare chart data
    attendance_data = {
        'dates': dates,
        'present': present_data,
        'absent': absent_data,
        'halfday': halfday_data,
        'leave': leave_data
    }

    # Calculate last recorded date
    if year > current_date.year or (year == current_date.year and month > current_date.month):
        last_recorded_date = "N/A"
    elif year == current_date.year and month == current_date.month:
        last_recorded_date = "29/05/2025"
    else:
        last_recorded_date = f"{total_days:02d}/{month:02d}/{year}"

    # Get leave balance for the selected month
    leaves_allowed, leaves_taken = get_leave_balance(company_id, name, year_month)

    logger.debug(f"Attendance data for {name} in {year_month}: Present: {sum(1 for x in present_data if x > 0)}, "
                 f"Absent: {sum(1 for x in absent_data if x > 0)}, "
                 f"HalfDay: {sum(1 for x in halfday_data if x > 0)}, "
                 f"Leave: {sum(1 for x in leave_data if x > 0)}, "
                 f"Total Days: {total_days}")

    return render_template(
        'student_dashboard.html',
        company_id=company_id,
        company_name=company_name,
        student_name=name,
        attendance_records=attendance_records,
        attendance_data=attendance_data,
        present_days=present_days,
        total_days=total_days,
        leaves_taken=leaves_taken,
        leaves_allowed=leaves_allowed,
        month_name=month_name,
        month=month,
        year=year,
        last_recorded_date=last_recorded_date
    )
def get_leave_balance(company_id, user_name, year_month):
    """Fetch or initialize leave balance for a user in a given month."""
    result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Leaves!A2:F').execute()
    leave_records = result.get('values', [])

    # Get company's LeavesPerMonth
    company_result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:E').execute()
    companies = company_result.get('values', [])
    company = next((c for c in companies if c[0] == company_id), None)
    leaves_per_month = int(company[4]) if company and len(company) > 4 and company[4] else 2  # Default to 2

    current_record = None
    for record in leave_records:
        if (len(record) >= 3 and record[0] == company_id and 
            record[1] == user_name and record[2] == year_month):
            current_record = record
            break
    
    if not current_record:
        # Get previous month's record
        year, month = map(int, year_month.split('-'))
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        prev_year_month = f"{prev_year}-{prev_month:02d}"
        
        carried_leaves = 0
        prev_record = None
        for record in leave_records:
            if (len(record) >= 6 and record[0] == company_id and 
                record[1] == user_name and record[2] == prev_year_month):
                prev_record = record
                break
        
        if prev_record:
            leaves_allowed = int(prev_record[3]) if len(prev_record) > 3 else leaves_per_month
            leaves_taken = int(prev_record[4]) if len(prev_record) > 4 else 0
            carried_leaves = max(0, leaves_allowed - leaves_taken)
        
        leaves_allowed = leaves_per_month + carried_leaves
        leaves_taken = 0
        new_record = [company_id, user_name, year_month, str(leaves_allowed), '0', str(carried_leaves)]
        leave_records.append(new_record)
        service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range='Leaves!A2:F',
            valueInputOption='RAW',
            body={'values': leave_records}
        ).execute()
        logger.info(f"Initialized leave record for {user_name} in {year_month}: {leaves_allowed} allowed")
        return leaves_allowed, 0
    
    leaves_allowed = int(current_record[3]) if len(current_record) > 3 else leaves_per_month
    leaves_taken = int(current_record[4]) if len(current_record) > 4 else 0
    logger.debug(f"Leave balance for {user_name} in {year_month}: {leaves_allowed} allowed, {leaves_taken} taken")
    return leaves_allowed, leaves_taken


def update_leave_balance(company_id, name, year_month, status):
    """Update leave balance when an absence is recorded."""
    if status != 'Absent':
        return
    
    result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Leaves!A2:F').execute()
    leave_records = result.get('values', [])
    
    for i, record in enumerate(leave_records):
        if (len(record) >= 3 and record[0] == company_id and record[1] == name and record[2] == year_month):
            leaves_taken = int(record[5]) if len(record) > 5 else 0
            leaves_allowed = int(record[4]) if len(record) > 4 else 2
            if leaves_taken < leaves_allowed:
                record[5] = str(leaves_taken + 1)
                leave_records[i] = record
                service.spreadsheets().values().update(
                    spreadsheetId=SHEET_ID,
                    range='Leaves!A2:F',
                    valueInputOption='RAW',
                    body={'values': leave_records}
                ).execute()
            break
    else:
        leaves_allowed, _ = get_leave_balance(company_id, name, year_month)
        new_record = [company_id, name, year_month, str(leaves_allowed), '1', '0']
        leave_records.append(new_record)
        service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range='Leaves!A2:F',
            valueInputOption='RAW',
            body={'values': leave_records}
        ).execute()


@app.route('/update_leave_config', methods=['POST'])
def update_leave_config():
    print(f'{request.form}')
    if 'username' not in session or session.get('role') != 'company_admin':
        return redirect(url_for('login'))

    company_id = request.form.get('company_id')
    leaves_per_month = request.form.get('leaves_per_month')
    

    if not leaves_per_month.isdigit() or int(leaves_per_month) < 0:
        flash('Invalid number of leaves', 'error')
        return redirect(url_for('admin_panel'))

    try:
        # Fetch Companies sheet
        result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:E').execute()
        companies = result.get('values', [])
        logger.debug(f"Fetched {len(companies)} companies from sheet")

        # Find and update the company
        for i, company in enumerate(companies):
            if company[1] == company_id:
                # Ensure company has 5 columns
                while len(company) < 5:
                    company.append('2')  # Default LeavesPerMonth
                    logger.info(f"Padded company {company_id} with default LeavesPerMonth=2")
                company[4] = leaves_per_month
                companies[i] = company
                logger.info(f"Updated LeavesPerMonth to {leaves_per_month} for company {company_id}")
                break
        else:
            logger.error(f"Company {company_id} not found in Companies sheet")
            flash('Company not found', 'error')
            return redirect(url_for('admin_panel'))

        # Update the sheet
        service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range='Companies!A2:E',
            valueInputOption='RAW',
            body={'values': companies}
        ).execute()
        logger.debug("Companies sheet updated successfully")

        flash('Leave configuration updated successfully', 'success')
    except Exception as e:
        logger.error(f"Error updating leave configuration: {str(e)}")
        flash(f'Error updating leave configuration: {str(e)}', 'error')
        return redirect(url_for('admin_panel'))

    return redirect(url_for('admin_panel'))


# @app.route('/admin_panel', methods=['GET', 'POST'])
# def admin_panel():
#     if session.get('role') not in ['company_admin']:
#         return redirect(url_for('login'))

#     company_id = session.get('company_id')
#     company_name = session.get('company_name', 'Unknown Company')
#     encodings = load_encodings(company_id)
#     names = list(encodings.keys())

#     try:
#         safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A1'))
#     except RetryError as retry_err:
#         underlying_exc = retry_err.last_attempt.exception()
#         if isinstance(underlying_exc, HttpError) and underlying_exc.resp.status in [500, 503]:
#             logger.error(f"Failed to check company sheet {company_id} after retries: {underlying_exc}")
#             flash("Google Sheets API is unavailable. Please try again later.", "error")
#             return redirect(url_for('logout'))
#         else:
#             # Could be a 404 or something else – handle sheet creation
#             safe_api_call(
#                 service.spreadsheets().batchUpdate(
#                     spreadsheetId=SHEET_ID,
#                     body={'requests': [{'addSheet': {'properties': {'title': company_id}}}]}
#                 )
#             )
#             safe_api_call(
#                 service.spreadsheets().values().update(
#                     spreadsheetId=SHEET_ID,
#                     range=f'{company_id}!A1',
#                     valueInputOption='RAW',
#                     body={'values': [['Name']]}
#                 )
#             )

#     attendance = read_attendance_from_sheet(company_id)
#     today = datetime.now().strftime('%d/%m/%Y')
#     initial_data = {name: {
#         'checkin': '',
#         'checkout': '',
#         'status': 'Absent',
#         'allow_checkout': False,
#         'hours': '',
#         'checkin_image': None,
#         'day_status': 'N/A'
#     } for name in names}

#     todays_attendance = [r for r in attendance if r[1] == today]
#     present_count = 0
#     for name, date, checkin, checkout, status, expected_checkout, hours, day_status in todays_attendance:
#         checkin_formatted = checkin.split(':')[0] + ':' + checkin.split(':')[1] if checkin else ''
#         checkout_formatted = checkout.split(':')[0] + ':' + checkout.split(':')[1] if checkout else ''
#         initial_data[name] = {
#             'checkin': checkin_formatted,
#             'checkout': checkout_formatted,
#             'status': status,
#             'allow_checkout': bool(checkin and not checkout),
#             'hours': hours,
#             'checkin_image': get_checkin_image_base64(company_id, name, today),
#             'day_status': day_status if day_status else 'N/A'
#         }
#         if status == 'Present':
#             present_count += 1

#     absent_count = len(names) - present_count

#     if request.method == 'POST' and 'force_checkout' in request.form:
#         name = request.form['force_checkout']
#         now = datetime.now().strftime('%H:%M:%S')
#         for record in todays_attendance:
#             if record[0] == name and not record[3]:
#                 record[3] = now
#                 record[4] = 'Present'
#                 record[6] = calculate_hours(record[2], now)
#         try:
#             update_sheet(company_id, attendance)
#             image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{today.replace('/', '-')}.jpg")
#             if os.path.exists(image_path):
#                 os.remove(image_path)
#             flash("Checkout forced successfully.", "success")
#         except Exception as e:
#             flash("Network error: Could not update attendance.", "error")
#         return redirect(url_for('admin_panel'))

#     # Add current month and year for the form
#     current_date = datetime.now()
#     current_month = current_date.month
#     current_year = current_date.year

#     return render_template('admin_panel.html',
#                            names=names,
#                            attendance=todays_attendance,
#                            initial_data=initial_data,
#                            today=today,
#                            company_name=company_name,
#                            present_count=present_count,
#                            absent_count=absent_count,
#                            show_full_dashboard=len(names) > 0,
#                            current_month=current_month,
#                            current_year=current_year)

@app.route('/admin_panel', methods=['GET', 'POST'])
def admin_panel():
    if 'username' not in session or session.get('role') != 'company_admin':
        return redirect(url_for('login'))

    company_id = session.get('company_id')
    company_name = session.get('company_name')
    today = datetime.now().strftime('%d/%m/%Y')
    current_year_month = datetime.now().strftime('%Y-%m')

    # Fetch company leave configuration
    result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:E').execute()
    companies = result.get('values', [])
    company = next((c for c in companies if c[0] == company_id), None)
    leaves_per_month = int(company[4]) if company and len(company) > 4 and company[4] else 2
    logger.debug(f"LeavesPerMonth for {company_id}: {leaves_per_month}")

    # Fetch users
    result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A2:B').execute()
    users = result.get('values', [])
    names = [user[1] for user in users if user[0] == company_id]

    # Fetch leave balances
    leave_balances = []
    leave_result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Leaves!A2:F').execute()
    leave_records = leave_result.get('values', [])
    for record in leave_records:
        if (len(record) >= 3 and record[0] == company_id and record[2] == current_year_month):
            leave_balances.append({
                'user_name': record[1],
                'leaves_allowed': int(record[3]) if len(record) > 3 else leaves_per_month,
                'leaves_taken': int(record[4]) if len(record) > 4 else 0,
                'leaves_carried': int(record[5]) if len(record) > 5 else 0
            })

    # Fetch attendance and generate absent records
    attendance = read_attendance_from_sheet(company_id)
    initial_data = {}
    for name in names:
        checkin, checkout, status, hours, checkin_image, day_status = '', '', 'Absent', '0', '', 'Absent'
        allow_checkout = False
        today_records = [r for r in attendance if r[0] == name and r[1] == today]
        if today_records:
            record = today_records[-1]
            checkin = record[2] or ''
            checkout = record[3] or ''
            status = record[4] or 'Absent'
            hours = record[6] or '0'
            checkin_image = get_checkin_image_base64(company_id, name, today) or ''
            day_status = record[7] if len(record) > 7 else 'Absent'
            allow_checkout = checkin and not checkout
        initial_data[name] = {
            'checkin': checkin,
            'checkout': checkout,
            'status': status,
            'hours': hours,
            'checkin_image': checkin_image,
            'day_status': day_status,
            'allow_checkout': allow_checkout
        }

    if request.method == 'POST':
        if 'force_checkout' in request.form:
            name = request.form['force_checkout']
            now = datetime.now()
            time_str = now.strftime('%H:%M:%S')
            attendance = read_attendance_from_sheet(company_id)
            for record in attendance:
                if record[0] == name and record[1] == today and record[2] and not record[3]:
                    record[3] = time_str
                    record[4] = 'Present'
                    record[6] = calculate_hours(record[2], time_str)
                    record[7] = 'Present'  # Update day_status
                    update_sheet(company_id, attendance)
                    flash('Checkout forced successfully', 'success')
                    break
            return redirect(url_for('admin_panel'))

    return render_template(
        'admin_panel.html',
        company_id=company_id,
        company_name=company_name,
        today=today,
        current_year_month=current_year_month,
        names=names,
        initial_data=initial_data,
        show_full_dashboard=True,
        leaves_per_month=leaves_per_month,
        leave_balances=leave_balances
    )


# @app.route('/add_user', methods=['GET', 'POST'])
# def add_user():
#     if session.get('role') != 'company_admin':
#         return redirect(url_for('login'))

#     company_id = session.get('company_id')
#     if request.method == 'POST':
#         try:
#             if not request.is_json:
#                 return jsonify({'error': 'Invalid request. JSON data required.'}), 400

#             data = request.get_json()
#             name = data.get('name').strip()
#             images = data.get('images', [])

#             if not name:
#                 return jsonify({'error': 'Name is required.'}), 400
#             if not images:
#                 return jsonify({'error': 'No images provided.'}), 400

#             encodings = load_encodings(company_id)
#             if name in encodings:
#                 return jsonify({'error': 'User already registered. Try a different name.'}), 400

#             person_dir = os.path.join(IMAGES_DIR, company_id, name)
#             if not os.path.exists(person_dir):
#                 os.makedirs(person_dir)

#             known_face_encodings = []
#             image_count = 0
#             for image_data in images:
#                 try:
#                     if ',' in image_data:
#                         image_data = image_data.split(',')[1]
#                     image_bytes = base64.b64decode(image_data)
#                     image = Image.open(io.BytesIO(image_bytes))
#                     frame = np.array(image)
#                     if frame.shape[-1] == 4:
#                         frame = frame[:, :, :3]
#                     rgb_frame = frame
#                     face_locations = face_recognition.face_locations(rgb_frame)
#                     if len(face_locations) != 1:
#                         continue
#                     face_encoding = face_recognition.face_encodings(rgb_frame, face_locations)[0]
#                     if face_encoding.shape != (128,):
#                         continue
#                     known_face_encodings.append(face_encoding)
#                     image_count += 1
#                     image_path = os.path.join(person_dir, f'{name}_{image_count}.jpg')
#                     Image.fromarray(frame).save(image_path)
#                 except Exception as e:
#                     logger.error(f"Error processing image {image_count + 1}: {str(e)}")
#                     continue

#             if known_face_encodings:
#                 if len(known_face_encodings) < 5:
#                     return jsonify({'error': 'Insufficient face captures (less than 5). Try again with better lighting or more angles.'}), 400
#                 encodings[name] = known_face_encodings
#                 save_encodings(company_id, encodings)
#                 safe_api_call(
#                     service.spreadsheets().values().append(
#                         spreadsheetId=SHEET_ID,
#                         range='Users!A2:B2',
#                         valueInputOption='RAW',
#                         body={'values': [[company_id, name]]}
#                     )
#                 )
#                 headers = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A1:1')).get('values', [[]])[0]
#                 all_data = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A2:B')).get('values', [])
#                 all_data.append([name])
#                 safe_api_call(
#                     service.spreadsheets().values().update(
#                         spreadsheetId=SHEET_ID,
#                         range=f'{company_id}!A1',
#                         valueInputOption='RAW',
#                         body={'values': [headers]}
#                     )
#                 )
#                 if all_data:
#                     safe_api_call(
#                         service.spreadsheets().values().update(
#                             spreadsheetId=SHEET_ID,
#                             range=f'{company_id}!A2',
#                             valueInputOption='RAW',
#                             body={'values': all_data}
#                         )
#                     )
#                 return jsonify({'success': f'User {name} registered successfully'}), 200
#             else:
#                 return jsonify({'error': 'No valid faces detected.'}), 400

#         except Exception as e:
#             logger.error(f"Error in add_user: {str(e)}")
#             return jsonify({'error': f'Error: {str(e)}'}), 500

#     return render_template('add_user.html', error=None)

india_offset = timedelta(hours=5, minutes=30)

@app.route('/add_user', methods=['GET', 'POST'])
def add_user():
    if session.get('role') != 'company_admin':
        return redirect(url_for('login'))

    company_id = session.get('company_id')
    if request.method == 'POST':
        try:
            if not request.is_json:
                return jsonify({'error': 'Invalid request. JSON data required.'}), 400

            data = request.get_json()
            name = data.get('name').strip()
            images = data.get('images', [])

            if not name:
                return jsonify({'error': 'Name is required.'}), 400
            if not images:
                return jsonify({'error': 'No images provided.'}), 400

            encodings = load_encodings(company_id)
            if name in encodings:
                return jsonify({'error': 'User already registered. Try a different name.'}), 400

            person_dir = os.path.join(IMAGES_DIR, company_id, name)
            if not os.path.exists(person_dir):
                os.makedirs(person_dir)

            known_face_encodings = []
            image_paths = []  # Track image paths for deletion
            image_count = 0
            for image_data in images:
                try:
                    if ',' in image_data:
                        image_data = image_data.split(',')[1]
                    image_bytes = base64.b64decode(image_data)
                    image = Image.open(io.BytesIO(image_bytes))
                    frame = np.array(image)
                    if frame.shape[-1] == 4:
                        frame = frame[:, :, :3]
                    rgb_frame = frame
                    face_locations = face_recognition.face_locations(rgb_frame)
                    if len(face_locations) != 1:
                        continue
                    face_encoding = face_recognition.face_encodings(rgb_frame, face_locations)[0]
                    if face_encoding.shape != (128,):
                        continue
                    known_face_encodings.append(face_encoding)
                    image_count += 1
                    image_path = os.path.join(person_dir, f'{name}_{image_count}.jpg')
                    Image.fromarray(frame).save(image_path)
                    image_paths.append(image_path)  # Store path for later deletion
                    logger.debug(f"Saved image: {image_path}")
                except Exception as e:
                    logger.error(f"Error processing image {image_count + 1}: {str(e)}")
                    continue

            if known_face_encodings:
                if len(known_face_encodings) < 5:
                    return jsonify({'error': 'Insufficient face captures (less than 5). Try again with better lighting or more angles.'}), 400
                encodings[name] = known_face_encodings
                save_encodings(company_id, encodings)
                logger.info(f"Face encodings saved for user {name} in company {company_id}")

                # Delete temporary images
                for img_path in image_paths:
                    try:
                        if os.path.exists(img_path):
                            os.remove(img_path)
                            logger.debug(f"Deleted image: {img_path}")
                        else:
                            logger.warning(f"Image not found for deletion: {img_path}")
                    except Exception as e:
                        logger.error(f"Failed to delete image {img_path}: {str(e)}")
                        # Continue to avoid affecting response

                # Append user to Users sheet
                safe_api_call(
                    service.spreadsheets().values().append(
                        spreadsheetId=SHEET_ID,
                        range='Users!A2:B2',
                        valueInputOption='RAW',
                        body={'values': [[company_id, name]]}
                    )
                )

                # Check if company sheet exists
                try:
                    spreadsheet = safe_api_call(service.spreadsheets().get(spreadsheetId=SHEET_ID))
                    sheet_exists = any(sheet['properties']['title'] == company_id for sheet in spreadsheet.get('sheets', []))
                except HttpError as e:
                    logger.error(f"Error checking if sheet {company_id} exists: {str(e)}")
                    return jsonify({'error': f'Failed to verify company sheet: {str(e)}'}), 500

                # Create company sheet if it doesn't exist
                if not sheet_exists:
                    try:
                        requests = [{
                            'addSheet': {
                                'properties': {
                                    'title': company_id,
                                    'gridProperties': {'rowCount': 1000, 'columnCount': 20}
                                }
                            }
                        }]
                        safe_api_call(service.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={'requests': requests}))
                        # Initialize headers
                        default_headers = ['Name', 'Date', 'Time Range', 'Status', 'Hours']
                        safe_api_call(
                            service.spreadsheets().values().update(
                                spreadsheetId=SHEET_ID,
                                range=f'{company_id}!A1:E1',
                                valueInputOption='RAW',
                                body={'values': [default_headers]}
                            )
                        )
                    except HttpError as e:
                        logger.error(f"Error creating sheet {company_id}: {str(e)}")
                        return jsonify({'error': f'Failed to create company sheet: {str(e)}'}), 500

                # Update company-specific sheet
                try:
                    headers = safe_api_call(
                        service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A1:1')
                    ).get('values', [[]])[0] or ['Name']
                    all_data = safe_api_call(
                        service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A2:B')
                    ).get('values', [])
                    all_data.append([name])
                    safe_api_call(
                        service.spreadsheets().values().update(
                            spreadsheetId=SHEET_ID,
                            range=f'{company_id}!A1',
                            valueInputOption='RAW',
                            body={'values': [headers]}
                        )
                    )
                    if all_data:
                        safe_api_call(
                            service.spreadsheets().values().update(
                                spreadsheetId=SHEET_ID,
                                range=f'{company_id}!A2',
                                valueInputOption='RAW',
                                body={'values': all_data}
                            )
                        )
                except HttpError as e:
                    logger.warning(f"Non-critical error updating company sheet {company_id}: {str(e)}")
                    # Continue despite error, as user is already registered
                    pass

                return jsonify({'success': f'User {name} registered successfully'}), 200
            else:
                return jsonify({'error': 'No valid faces detected.'}), 400

        except Exception as e:
            logger.error(f"Error in add_user: {str(e)}")
            return jsonify({'error': f'Error: {str(e)}'}), 500

    return render_template('add_user.html', error=None)


@app.route('/delete_user', methods=['GET'])
def delete_user():
    if session.get('role') != 'company_admin':
        return redirect(url_for('login'))

    company_id = session.get('company_id')
    name = request.args.get('name', '')

    if not name:
        flash("No user specified for deletion.", "error")
        return redirect(url_for('admin_panel'))

    encodings = load_encodings(company_id)
    if name in encodings:
        try:
            del encodings[name]
            save_encodings(company_id, encodings)
            user_data = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Users!A2:B'))
            user_values = user_data.get('values', [])
            updated_users = [row for row in user_values if row[1] != name or row[0] != company_id]
            safe_api_call(
                service.spreadsheets().values().update(
                    spreadsheetId=SHEET_ID,
                    range='Users!A2:B',
                    valueInputOption='RAW',
                    body={'values': updated_users}
                )
            )
            all_data = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A2:B'))
            updated_data = [row for row in all_data.get('values', []) if row[0] != name]
            headers = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f'{company_id}!A1:C')).get('values', [[]])[0]
            safe_api_call(
                service.spreadsheets().values().update(
                    spreadsheetId=SHEET_ID,
                    range=f'{company_id}!A1',
                    valueInputOption='RAW',
                    body={'values': [headers]}
                )
            )
            if updated_data:
                safe_api_call(
                    service.spreadsheets().values().update(
                        spreadsheetId=SHEET_ID,
                        range=f'{company_id}!A2',
                        valueInputOption='RAW',
                        body={'values': updated_data}
                    )
                )
            person_dir = os.path.join(IMAGES_DIR, company_id, name)
            if os.path.exists(person_dir):
                shutil.rmtree(person_dir)
            for image_file in os.listdir(TEMP_CHECKIN_IMAGES_DIR):
                if image_file.startswith(f"{company_id}_{name}_"):
                    os.remove(os.path.join(TEMP_CHECKIN_IMAGES_DIR, image_file))
            flash(f"User {name} deleted successfully", "success")
            return redirect(url_for('admin_panel'))
        except Exception as e:
            logger.error(f"Error deleting user {name}: {str(e)}")
            flash(f"Error deleting user {name}: {str(e)}", "error")
            return redirect(url_for('admin_panel'))
    flash(f"User {name} not found.", "error")
    return redirect(url_for('admin_panel'))

# @app.route('/user_panel', methods=['GET', 'POST'])
# def user_panel():
#     if session.get('role') == 'super_admin':
#         return redirect(url_for('super_admin'))

#     company_id = session.get('company_id')
#     known_faces = load_encodings(company_id)
#     action = "Welcome, please start recognition"
#     name = "Unknown"

#     if request.method == 'POST':
#         try:
#             # Check if the request has the correct Content-Type
#             if not request.is_json:
#                 logger.error("Invalid request: JSON data required")
#                 return jsonify({'action': 'Invalid request. Content-Type must be application/json.', 'name': name}), 400

#             # Parse JSON data
#             data = request.get_json(silent=True)
#             if data is None:
#                 logger.error("Invalid JSON format in request")
#                 return jsonify({'action': 'Invalid JSON format. Please ensure the request body is valid JSON.', 'name': name}), 400

#             # Check for the 'image' field
#             if 'image' not in data:
#                 logger.error("No image provided in request")
#                 return jsonify({'action': 'No image provided. Please capture an image.', 'name': name}), 400

#             image_data = data['image']
#             if not isinstance(image_data, str):
#                 logger.error("Image data is not a string")
#                 return jsonify({'action': 'Image data must be a string.', 'name': name}), 400

#             if not image_data:
#                 logger.error("Image data is empty")
#                 return jsonify({'action': 'Image data is empty. Please capture a valid image.', 'name': name}), 400

#             logger.debug("Received base64 image data")

#             # Step 1: Decode base64 image
#             try:
#                 # Remove the data URI prefix if present (e.g., "data:image/jpeg;base64,")
#                 if ',' in image_data:
#                     image_data = image_data.split(',')[1]
#                 # Ensure proper padding for base64 decoding
#                 image_data += '=' * (-len(image_data) % 4)
#                 image_bytes = base64.b64decode(image_data, validate=True)
#                 logger.debug("Base64 decoded successfully")
#             except (base64.binascii.Error, ValueError) as e:
#                 logger.error(f"Base64 decoding error: {str(e)}")
#                 return jsonify({'action': f'Invalid base64 image data: {str(e)}', 'name': name}), 400

#             # Step 2: Convert image to RGB format for face_recognition
#             try:
#                 image = Image.open(io.BytesIO(image_bytes))
#                 if image.mode != 'RGB':
#                     image = image.convert('RGB')
#                 frame = np.array(image)
#                 rgb_frame = frame
#                 logger.debug(f"Image converted to RGB, shape: {frame.shape}")
#             except Exception as e:
#                 logger.error(f"Error opening or converting image: {str(e)}")
#                 return jsonify({'action': 'Error processing image. Ensure the image is a valid JPEG.', 'name': name}), 400

#             # Step 3: Detect faces in the image
#             try:
#                 face_locations = face_recognition.face_locations(rgb_frame, model="hog")
#                 if not face_locations:
#                     logger.warning("No faces detected in image")
#                     return jsonify({'action': 'No face detected. Ensure a face is clearly visible in the frame.', 'name': name}), 400
#                 logger.debug(f"Detected {len(face_locations)} face(s)")
#             except Exception as e:
#                 logger.error(f"Error detecting faces: {str(e)}")
#                 return jsonify({'action': 'Error detecting faces. Ensure the image quality is good and try again.', 'name': name}), 500

#             # Step 4: Generate face encodings
#             try:
#                 face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
#                 if not face_encodings:
#                     logger.warning("No face encodings generated")
#                     return jsonify({'action': 'Face encoding failed. Ensure the face is clear and well-lit.', 'name': name}), 400
#                 logger.debug(f"Generated {len(face_encodings)} face encoding(s)")
#             except Exception as e:
#                 logger.error(f"Error generating face encodings: {str(e)}")
#                 return jsonify({'action': 'Error encoding face. Ensure the image quality is good and try again.', 'name': name}), 500

#             # Step 5: Match the face with known faces
#             for face_encoding in face_encodings:
#                 try:
#                     match_result = find_best_match(face_encoding, known_faces)
#                     if match_result:
#                         name, best_distance = match_result
#                         logger.info(f"Recognized: {name} with distance {best_distance}")
#                         attendance = read_attendance_from_sheet(company_id)
#                         today = datetime.now().strftime('%d/%m/%Y')
#                         today_records = [r for r in attendance if r[0] == name and r[1] == today]
#                         now = datetime.now()

#                         if not today_records:
#                             noon = datetime.strptime(f"{today} 12:00:00", '%d/%m/%Y %H:%M:%S')
#                             if now > noon:
#                                 action = "Check-in not allowed after 12:00 PM."
#                             else:
#                                 success, expected_checkout, day_status = log_attendance(company_id, name, 'checkin')
#                                 if success:
#                                     action = f"Checked in successfully. Expected check-out: {expected_checkout}. Day Status: {day_status}"
#                                     image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{today.replace('/', '-')}.jpg")
#                                     try:
#                                         image.save(image_path)
#                                         logger.info(f"Saved check-in image: {name}")
#                                     except Exception as e:
#                                         logger.error(f"Error saving check-in image: {str(e)}")
#                                         action += " (Warning: Failed to save check-in image)"
#                                 else:
#                                     action = "Error processing check-in."
#                         else:
#                             last_record = today_records[-1]
#                             checkin_time = datetime.strptime(last_record[2], '%H:%M:%S') if last_record[2] else None
#                             checkout_time = datetime.strptime(last_record[3], '%H:%M:%S') if last_record[3] else None

#                             if checkout_time:
#                                 action = 'Attendance completed for today'
#                             elif checkin_time and not checkout_time:
#                                 time_since_checkin = now - datetime.combine(date.today(), checkin_time.time())
#                                 expected_checkout = datetime.strptime(last_record[5], '%H:%M:%S') if last_record[5] else datetime.strptime('18:30:00', '%H:%M:%S')
#                                 if time_since_checkin >= timedelta(hours=7):
#                                     success, hours, day_status = log_attendance(company_id, name, 'checkout')
#                                     if success:
#                                         action = f"Checked out successfully. Hours: {hours}. Day Status: {day_status}"
#                                         image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{hours}_{today.replace('/', '-')}.jpg")
#                                         if os.path.exists(image_path):
#                                             try:
#                                                 os.remove(image_path)
#                                                 logger.info(f"Deleted check-in image: {name}")
#                                             except Exception as e:
#                                                 logger.error(f"Error deleting check-in image: {str(e)}")
#                                     else:
#                                         action = "Error processing check-out."
#                                 else:
#                                     time_to_checkout = datetime.combine(date.today(), expected_checkout.time()) - now
#                                     if time_to_checkout.total_seconds() > 0:
#                                         hours, remainder = divmod(time_to_checkout.total_seconds(), 3600)
#                                         minutes, _ = divmod(remainder, 60)
#                                         action = f"Cannot check out yet, wait until {last_record[5]} ({int(hours)}h {int(minutes)}m)"
#                                     else:
#                                         action = f"Cannot check out yet, minimum 7 hours required."
#                             else:
#                                 action = 'Invalid attendance state'
#                     else:
#                         name = "Unknown"
#                         action = "Unknown user."
#                 except Exception as e:
#                     logger.error(f"Error matching face: {str(e)}")
#                     return jsonify({'action': 'Error matching face. Ensure the face matches a registered user.', 'name': name}), 500

#             return jsonify({'action': action, 'name': name})

#         except Exception as e:
#             logger.error(f"Unexpected error in user_panel: {str(e)}")
#             return jsonify({'action': f'Unexpected server error: {str(e)}', 'name': name}), 500

#     return render_template('user_panel.html', name=name, action=action, known_faces=known_faces)


# @app.route('/user_panel', methods=['GET', 'POST'])
# def user_panel():
#     if session.get('role') == 'super_admin':
#         return redirect(url_for('super_admin'))

#     company_id = session.get('company_id')
#     known_faces = load_encodings(company_id)
#     action = "Welcome, please start recognition"
#     name = "Unknown"

#     if request.method == 'POST':
#         try:
#             # Check if the request has the correct Content-Type
#             if not request.is_json:
#                 logger.error("Invalid request: JSON data required")
#                 return jsonify({'action': 'Invalid request. Content-Type must be application/json.', 'name': name}), 400

#             # Parse JSON data
#             data = request.get_json(silent=True)
#             if data is None:
#                 logger.error("Invalid JSON format in request")
#                 return jsonify({'action': 'Invalid JSON format. Please ensure the request body is valid JSON.', 'name': name}), 400

#             # Check for the 'image' field
#             if 'image' not in data:
#                 logger.error("No image provided in request")
#                 return jsonify({'action': 'No image provided. Please capture an image.', 'name': name}), 400

#             image_data = data['image']
#             if not isinstance(image_data, str):
#                 logger.error("Image data is not a string")
#                 return jsonify({'action': 'Image data must be a string.', 'name': name}), 400

#             if not image_data:
#                 logger.error("Image data is empty")
#                 return jsonify({'action': 'Image data is empty. Please capture a valid image.', 'name': name}), 400

#             logger.debug("Received base64 image data")

#             # Step 1: Decode base64 image
#             try:
#                 if ',' in image_data:
#                     image_data = image_data.split(',')[1]
#                 image_data += '=' * (-len(image_data) % 4)
#                 image_bytes = base64.b64decode(image_data, validate=True)
#                 logger.debug("Base64 decoded successfully")
#             except (base64.binascii.Error, ValueError) as e:
#                 logger.error(f"Base64 decoding error: {str(e)}")
#                 return jsonify({'action': f'Invalid base64 image data: {str(e)}', 'name': name}), 400

#             # Step 2: Convert image to RGB format for face_recognition
#             try:
#                 image = Image.open(io.BytesIO(image_bytes))
#                 if image.mode != 'RGB':
#                     image = image.convert('RGB')
#                 frame = np.array(image)
#                 rgb_frame = frame
#                 logger.debug(f"Image converted to RGB, shape: {frame.shape}")
#             except Exception as e:
#                 logger.error(f"Error opening or converting image: {str(e)}")
#                 return jsonify({'action': 'Error processing image. Ensure the image is a valid JPEG.', 'name': name}), 400

#             # Step 3: Detect faces in the image
#             try:
#                 face_locations = face_recognition.face_locations(rgb_frame, model="hog")
#                 if not face_locations:
#                     logger.warning("No faces detected in image")
#                     return jsonify({'action': 'No face detected. Ensure a face is clearly visible in the frame.', 'name': name}), 400
#                 logger.debug(f"Detected {len(face_locations)} face(s)")
#             except Exception as e:
#                 logger.error(f"Error detecting faces: {str(e)}")
#                 return jsonify({'action': 'Error detecting faces. Ensure the image quality is good and try again.', 'name': name}), 500

#             # Step 4: Generate face encodings
#             try:
#                 face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
#                 if not face_encodings:
#                     logger.warning("No face encodings generated")
#                     return jsonify({'action': 'Face encoding failed. Ensure the face is clear and well-lit.', 'name': name}), 400
#                 logger.debug(f"Generated {len(face_encodings)} face encoding(s)")
#             except Exception as e:
#                 logger.error(f"Error generating face encodings: {str(e)}")
#                 return jsonify({'action': 'Error encoding face. Ensure the image quality is good and try again.', 'name': name}), 500

#             # Step 5: Match the face with known faces
#             for face_encoding in face_encodings:
#                 try:
#                     match_result = find_best_match(face_encoding, known_faces)
#                     if match_result:
#                         name, best_distance = match_result
#                         logger.info(f"Recognized: {name} with distance {best_distance}")
#                         attendance = read_attendance_from_sheet(company_id)
#                         # Use 5:30 delta time instead of timezone
#                         india_offset = timedelta(hours=5, minutes=30)
#                         current_time = datetime.now() + india_offset
#                         today = current_time.strftime('%d/%m/%Y')
#                         time_str = current_time.strftime('%H:%M:%S')
#                         today_records = [r for r in attendance if r[0] == name and r[1] == today]
#                         now = current_time
                        
#                         print(f"[TIME] User panel accessed at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                         print(f"[DEBUG] Processing face recognition for {name} on {today} at {time_str}")

#                         if not today_records:
#                             # Get company-specific check-in time
#                             check_in_time, check_out_time, required_hours = get_company_times(company_id)
#                             if check_in_time is None:
#                                 logger.error("Company check-in time not set.")
#                                 return jsonify({'action': 'Error: Company check-in time not configured.', 'name': name}), 500

#                             # Calculate the check-in cutoff (check-in time + 2 hours)
#                             check_in_dt = datetime.strptime(f"{today} {check_in_time}", '%d/%m/%Y %H:%M:%S')
#                             check_in_cutoff = check_in_dt + timedelta(hours=2)
#                             # Apply india_offset to both check_in_dt and check_in_cutoff for comparison
#                             check_in_dt_with_offset = check_in_dt + india_offset
#                             check_in_cutoff_with_offset = check_in_cutoff + india_offset

#                             print(f"[TIME] Check-in window: {check_in_dt.strftime('%H:%M:%S')} to {check_in_cutoff.strftime('%H:%M:%S')} IST")
#                             print(f"[DEBUG] Current time: {now.strftime('%H:%M:%S')}, Cutoff: {check_in_cutoff_with_offset.strftime('%H:%M:%S')}")

#                             if now > check_in_cutoff_with_offset:
#                                 check_in_time_str = check_in_dt.strftime('%H:%M:%S')
#                                 check_in_cutoff_str = check_in_cutoff.strftime('%H:%M:%S')
#                                 action = f"Check-in not allowed after {check_in_cutoff_str} (2 hours past check-in time {check_in_time_str})."
#                                 print(f"[TIME] Check-in denied at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST - Too late")
#                             else:
#                                 print(f"[TIME] Processing check-in at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                                 success, expected_checkout, day_status = log_attendance(company_id, name, 'checkin')
#                                 if success:
#                                     action = f"Checked in successfully. Expected check-out: {expected_checkout}. Day Status: {day_status}"
#                                     print(f"[TIME] Check-in successful at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                                     image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{today.replace('/', '-')}.jpg")
#                                     try:
#                                         logger.debug(f"Saving check-in image to {image_path}")
#                                         image.save(image_path)
#                                         logger.info(f"Saved check-in image: {name}")
#                                         print(f"[TIME] Check-in image saved at: {current_time.strftime('%H:%M:%S')} IST")
#                                     except Exception as e:
#                                         logger.error(f"Error saving check-in image: {str(e)}")
#                                         action += " (Warning: Failed to save check-in image)"
#                                 else:
#                                     action = day_status if day_status == 'Check-in not allowed after 2 hours past check-in time. Admin must mark attendance.' else "Error processing check-in."
#                                     print(f"[TIME] Check-in failed at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                         else:
#                             last_record = today_records[-1]
#                             checkin_time = datetime.strptime(last_record[2], '%H:%M:%S') if last_record[2] else None
#                             checkout_time = datetime.strptime(last_record[3], '%H:%M:%S') if last_record[3] else None

#                             print(f"[DEBUG] Existing record found - Checkin: {last_record[2]}, Checkout: {last_record[3]}")

#                             if checkout_time:
#                                 action = 'Attendance completed for today'
#                                 print(f"[TIME] Attendance already completed at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                             elif checkin_time and not checkout_time:
#                                 # Use india_offset for time calculations
#                                 time_since_checkin = now - (datetime.combine(date.today(), checkin_time.time()) + india_offset)
#                                 expected_checkout = datetime.strptime(last_record[5], '%H:%M:%S') if last_record[5] else datetime.strptime('18:30:00', '%H:%M:%S')
                                
#                                 print(f"[DEBUG] Time since check-in: {time_since_checkin}, Required: 7 hours")
#                                 print(f"[TIME] Processing checkout attempt at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                                
#                                 if time_since_checkin >= timedelta(hours=7):
#                                     print(f"[TIME] Checkout allowed - processing at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                                     success, hours, day_status = log_attendance(company_id, name, 'checkout')
#                                     if success:
#                                         action = f"Checked out successfully. Hours: {hours}. Day Status: {day_status}"
#                                         print(f"[TIME] Check-out successful at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                                         image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{hours}_{today.replace('/', '-')}.jpg")
#                                         if os.path.exists(image_path):
#                                             try:
#                                                 os.remove(image_path)
#                                                 logger.info(f"Deleted check-in image: {name}")
#                                                 print(f"[TIME] Check-in image deleted at: {current_time.strftime('%H:%M:%S')} IST")
#                                             except Exception as e:
#                                                 logger.error(f"Error deleting check-in image: {str(e)}")
#                                     else:
#                                         action = "Error processing check-out."
#                                         print(f"[TIME] Check-out failed at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                                 else:
#                                     # Use india_offset for time calculations
#                                     time_to_checkout = (datetime.combine(date.today(), expected_checkout.time()) + india_offset) - now
#                                     if time_to_checkout.total_seconds() > 0:
#                                         hours, remainder = divmod(time_to_checkout.total_seconds(), 3600)
#                                         minutes, _ = divmod(remainder, 60)
#                                         action = f"Cannot check out yet, wait until {last_record[5]} ({int(hours)}h {int(minutes)}m)"
#                                         print(f"[TIME] Check-out denied at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST - Too early")
#                                     else:
#                                         action = f"Cannot check out yet, minimum 7 hours required."
#                                         print(f"[TIME] Check-out denied at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST - Insufficient hours")
#                             else:
#                                 action = 'Invalid attendance state'
#                                 print(f"[TIME] Invalid attendance state at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                     else:
#                         name = "Unknown"
#                         action = "Unknown user."
#                         print(f"[TIME] Unknown user detected at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
#                 except Exception as e:
#                     logger.error(f"Error matching face: {str(e)}")
#                     return jsonify({'action': 'Error matching face. Ensure the face matches a registered user.', 'name': name}), 500

#             return jsonify({'action': action, 'name': name})

#         except Exception as e:
#             logger.error(f"Unexpected error in user_panel: {str(e)}")
#             return jsonify({'action': f'Unexpected server error: {str(e)}', 'name': name}), 500

#     return render_template('user_panel.html', name=name, action=action, known_faces=known_faces)

@app.route('/user_panel', methods=['GET', 'POST'])
def user_panel():
    if session.get('role') == 'super_admin':
        return redirect(url_for('super_admin'))

    company_id = session.get('company_id')
    known_faces = load_encodings(company_id)
    action = "Welcome, please start recognition"
    name = "Unknown"

    india_offset = timedelta(hours=5, minutes=30)
    current_time = datetime.now() + india_offset

    if request.method == 'POST':
        try:
            # Check if the request has the correct Content-Type
            if not request.is_json:
                logger.error("Invalid request: JSON data required")
                return jsonify({'action': 'Invalid request. Content-Type must be application/json.', 'name': name}), 400

            # Parse JSON data
            data = request.get_json(silent=True)
            if data is None:
                logger.error("Invalid JSON format in request")
                return jsonify({'action': 'Invalid JSON format. Please ensure the request body is valid JSON.', 'name': name}), 400

            # Check for the 'image' field
            if 'image' not in data:
                logger.error("No image provided in request")
                return jsonify({'action': 'No image provided. Please capture an image.', 'name': name}), 400

            image_data = data['image']
            if not isinstance(image_data, str):
                logger.error("Image data is not a string")
                return jsonify({'action': 'Image data must be a string.', 'name': name}), 400

            if not image_data:
                logger.error("Image data is empty")
                return jsonify({'action': 'Image data is empty. Please capture a valid image.', 'name': name}), 400

            logger.debug("Received base64 image data")

            # Step 1: Decode base64 image
            try:
                if ',' in image_data:
                    image_data = image_data.split(',')[1]
                image_data += '=' * (-len(image_data) % 4)
                image_bytes = base64.b64decode(image_data, validate=True)
                logger.debug("Base64 decoded successfully")
            except (base64.binascii.Error, ValueError) as e:
                logger.error(f"Base64 decoding error: {str(e)}")
                return jsonify({'action': f'Invalid base64 image data: {str(e)}', 'name': name}), 400

            # Step 2: Convert image to RGB format for face_recognition
            try:
                image = Image.open(io.BytesIO(image_bytes))
                if image.mode != 'RGB':
                    image = image.convert('RGB')
                frame = np.array(image)
                rgb_frame = frame
                logger.debug(f"Image converted to RGB, shape: {frame.shape}")
            except Exception as e:
                logger.error(f"Error opening or converting image: {str(e)}")
                return jsonify({'action': 'Error processing image. Ensure the image is a valid JPEG.', 'name': name}), 400

            # Step 3: Detect faces in the image
            try:
                face_locations = face_recognition.face_locations(rgb_frame, model="hog")
                if not face_locations:
                    logger.warning("No faces detected in image")
                    return jsonify({'action': 'No face detected. Ensure a face is clearly visible in the frame.', 'name': name}), 400
                logger.debug(f"Detected {len(face_locations)} face(s)")
            except Exception as e:
                logger.error(f"Error detecting faces: {str(e)}")
                return jsonify({'action': 'Error detecting faces. Ensure the image quality is good and try again.', 'name': name}), 500

            # Step 4: Generate face encodings
            try:
                face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
                if not face_encodings:
                    logger.warning("No face encodings generated")
                    return jsonify({'action': 'Face encoding failed. Ensure the face is clear and well-lit.', 'name': name}), 400
                logger.debug(f"Generated {len(face_encodings)} face encoding(s)")
            except Exception as e:
                logger.error(f"Error generating face encodings: {str(e)}")
                return jsonify({'action': 'Error encoding face. Ensure the image quality is good and try again.', 'name': name}), 500

            # Step 5: Match the face with known faces
            for face_encoding in face_encodings:
                try:
                    match_result = find_best_match(face_encoding, known_faces)
                    if match_result:
                        name, best_distance = match_result
                        logger.info(f"Recognized: {name} with distance {best_distance}")
                        attendance = read_attendance_from_sheet(company_id)

                        # Use India Offset (5:30 hours)
                        # india_offset = timedelta(hours=5, minutes=30)
                        # current_time = datetime.now() + india_offset
                        today = current_time.strftime('%d/%m/%Y')
                        time_str = current_time.strftime('%H:%M:%S')
                        today_records = [r for r in attendance if r[0] == name and r[1] == today]
                        now = current_time

                        print(f"[TIME] User panel accessed at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                        print(f"[DEBUG] Processing face recognition for {name} on {today} at {time_str}")

                        if not today_records:
                            # No existing attendance record, process check-in
                            action = f"Check-in successful for {name} at {time_str}"
                            print(f"[TIME] Check-in successful at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                            success, expected_checkout, day_status = log_attendance(company_id, name, 'checkin')
                            
                            if success:
                                action = f"Checked in successfully. Expected check-out: {expected_checkout}. Day Status: {day_status}"
                                print(f"[TIME] Check-in image saved at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                                image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{today.replace('/', '-')}.jpg")
                                try:
                                    image.save(image_path)
                                    logger.info(f"Saved check-in image: {name}")
                                except Exception as e:
                                    logger.error(f"Error saving check-in image: {str(e)}")
                                    action += " (Warning: Failed to save check-in image)"
                            else:
                                action = "Error processing check-in."
                        else:
                            # Existing attendance record, process checkout or already completed
                            last_record = today_records[-1]
                            checkin_time = datetime.strptime(last_record[2], '%H:%M:%S') if last_record[2] else None
                            checkout_time = datetime.strptime(last_record[3], '%H:%M:%S') if last_record[3] else None

                            print(f"[DEBUG] Existing record found - Checkin: {last_record[2]}, Checkout: {last_record[3]}")

                            if checkout_time:
                                action = 'Attendance completed for today'
                                print(f"[TIME] Attendance already completed at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                            elif checkin_time and not checkout_time:
                                time_since_checkin = now - (datetime.combine(date.today(), checkin_time.time()) + india_offset)
                                if time_since_checkin >= timedelta(hours=7):
                                    print(f"[TIME] Checkout allowed - processing at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                                    success, hours, day_status = log_attendance(company_id, name, 'checkout')
                                    if success:
                                        action = f"Checked out successfully. Hours: {hours}. Day Status: {day_status}"
                                        print(f"[TIME] Check-out successful at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                                        image_path = os.path.join(TEMP_CHECKIN_IMAGES_DIR, f"{company_id}_{name}_{hours}_{today.replace('/', '-')}.jpg")
                                        if os.path.exists(image_path):
                                            try:
                                                os.remove(image_path)
                                                logger.info(f"Deleted check-in image: {name}")
                                                print(f"[TIME] Check-in image deleted at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                                            except Exception as e:
                                                logger.error(f"Error deleting check-in image: {str(e)}")
                                    else:
                                        action = "Error processing check-out."
                                        print(f"[TIME] Check-out failed at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                                else:
                                    time_to_checkout = datetime.combine(date.today(), datetime.strptime(last_record[5], '%H:%M:%S').time()) - now
                                    action = f"Cannot check out yet, minimum 7 hours required."
                                    print(f"[TIME] Check-out denied at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST - Insufficient hours")
                            else:
                                action = 'Invalid attendance state'
                                print(f"[TIME] Invalid attendance state at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                    else:
                        name = "Unknown"
                        action = "Unknown user."
                        print(f"[TIME] Unknown user detected at: {current_time.strftime('%Y-%m-%d %H:%M:%S')} IST")
                except Exception as e:
                    logger.error(f"Error matching face: {str(e)}")
                    return jsonify({'action': 'Error matching face. Ensure the face matches a registered user.', 'name': name}), 500

            return jsonify({'action': action, 'name': name})

        except Exception as e:
            logger.error(f"Unexpected error in user_panel: {str(e)}")
            return jsonify({'action': f'Unexpected server error: {str(e)}', 'name': name}), 500

    return render_template('user_panel.html', name=name, action=action, known_faces=known_faces)


import re

def get_company_times(company_id):
    try:
        result = service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:G').execute()
        companies = result.get('values', [])
        for company in companies:
            if company[0] == company_id:
                # Validate time format (HH:MM or HH:MM:SS)
                check_in = company[4] if len(company) > 4 and company[4] and re.match(r'^\d{2}:\d{2}(:\d{2})?$', company[4]) else None
                check_out = company[5] if len(company) > 5 and company[5] and re.match(r'^\d{2}:\d{2}(:\d{2})?$', company[5]) else None
                required_hours = float(company[6]) if len(company) > 6 and company[6] and company[6].replace('.', '', 1).isdigit() else None
                if check_in and check_out and required_hours:
                    return check_in, check_out, required_hours
                break
        logger.error(f"Company times not set for company_id: {company_id}")
        return None, None, None  # Return None values instead of redirect
    except Exception as e:
        logger.error(f"Error fetching company times: {e}")
        return None, None, None  # Return None values instead of redirect



@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    if session.get('role') != 'company_admin':
        return redirect(url_for('login'))

    company_id = session.get('company_id')
    name = request.form['mark_user']
    status = request.form[f'status_{name}']
    india = timezone('Asia/Kolkata')
    date_str = datetime.now().strftime('%d/%m/%Y')
    checkin = request.form.get(f'checkin_{name}', '')
    checkout = request.form.get(f'checkout_{name}', '')

    attendance = read_attendance_from_sheet(company_id)
    today_records = [r for r in attendance if r[0] == name and r[1] == date_str]
    if not today_records:
        try:
            day_status = 'Full Day'
            expected_checkout = '18:30:00'
            checkin_time = checkin + ':00' if checkin else None
            checkout_time = checkout + ':00' if checkout else None
            if checkin:
                checkin_dt = datetime.strptime(f'{datetime.today().strftime("%Y/%m-%d")} {checkin}', '%Y-%m-%d %H:%M')
                checkin_dt = checkin_dt.replace(tzinfo=timezone('Asia/Kolkata'))
                time_10_00 = datetime.combine(datetime.today(), time(10, 0))
                time_10_30 = datetime.combine(datetime.today(), time(10, 30))
                time_11_00 = datetime.combine(datetime.today(), time(11, 0))
                if time_10_00 <= checkin_dt < time_10_30:
                    expected_checkout = '18:30:00'
                elif time_10_30 <= checkin_dt <= time_11_00:
                    expected_checkout = (checkin_dt + timedelta(hours=8)).strftime('%H:%M:%S')
                elif checkin_dt > time_11_00:
                    expected_checkout = '18:30:00'
                    day_status = 'Half Day'
            hours = calculate_hours(checkin_time, checkout_time) if checkin_time and checkout_time else ''
            attendance.append([name, date_str, checkin_time, checkout_time, 'Present' if checkin_time else 'Absent', expected_checkout, hours, day_status])
            updated = True
        except ValueError:
            flash("Invalid time format.", "error")
            return redirect(url_for('admin_panel'))
    else:
        last_record = today_records[-1]
        updated = False
        if checkin and not last_record[2]:
            try:
                checkin_dt = datetime.strptime(checkin + ':00', '%H:%M:%S')
                time_10_00 = datetime.combine(datetime.today(), time(10, 0))
                time_10_30 = datetime.combine(datetime.today(), time(10, 30))
                time_11_00 = datetime.combine(datetime.today(), time(11, 0))
                day_status = 'Full Day'
                expected_checkout = '18:30:00'
                if time_10_00 <= checkin_dt < time_10_30:
                    expected_checkout = '18:30:00'
                elif time_10_30 <= checkin_dt <= time_11_00:
                    expected_checkout = (checkin_dt + timedelta(hours=8)).strftime('%H:%M:%S')
                elif checkin_dt > time_11_00:
                    expected_checkout = '18:30:00'
                    day_status = 'Half Day'
                last_record[2] = checkin + ':00'
                last_record[4] = 'Present'
                last_record[5] = expected_checkout
                last_record[7] = day_status
                updated = True
            except ValueError:
                flash("Invalid check-in time format.", "error")
                return redirect(url_for('admin_panel'))
        elif checkout and last_record[2] and not last_record[3]:
            try:
                checkin_time = datetime.strptime(last_record[2], '%H:%M:%S')
                checkout_dt = datetime.strptime(checkout + ':00', '%H:%M:%S')
                time_since_checkin = datetime.now(india) - datetime.combine(date.today(), checkin_time)
                expected_checkout = datetime.strptime(last_record[5], '%H:%M:%S') if last_record[5] else datetime.strptime('18:30:00', '%H:%M:%S')
                if time_since_checkin >= timedelta(hours=7) or session.get('role') == 'company_admin':
                    last_record[3] = checkout + ':00'
                    last_record[4] = 'Present'
                    last_record[6] = calculate_hours(last_record[2], last_record[3])
                    updated = True
                else:
                    flash("Checkout not allowed yet.", "error")
                    return redirect(url_for('admin_panel'))
            except ValueError:
                flash("Invalid check-out time format.", "error")
                return redirect(url_for('admin_panel'))

    if updated:
        try:
            update_sheet(company_id, attendance)
            flash("Attendance updated successfully.", "success")
        except Exception as e:
            logger.error(f"Error in mark_attendance: {str(e)}")
            flash("Error updating attendance.", "error")
            return redirect(url_for('admin_panel'))

    return redirect(url_for('admin_panel'))

def calculate_working_days(year, month):
    """Calculate the number of working days (Monday to Friday) in a given month."""
    first_day = datetime(year, month, 1)
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    last_day = (next_month - timedelta(days=1)).day

    working_days = []
    for day in range(1, last_day + 1):
        current_day = datetime(year, month, day)
        if current_day.weekday() < 5:  # Monday to Friday
            working_days.append(current_day.strftime('%d/%m/%Y'))
    return working_days

@app.route('/download_attendance')
def download_attendance():
    if session.get('role') != 'company_admin':
        return redirect(url_for('login'))

    company_id = session.get('company_id')
    attendance = read_attendance_from_sheet(company_id)
    all_names = list(load_encodings(company_id).keys())

    # Get the selected month from query parameter (format: MM-YYYY, e.g., 04-2025)
    selected_month = request.args.get('month')
    if not selected_month:
        # Default to current month if not specified
        selected_month = datetime.now().strftime('%m-%Y')
    
    try:
        month, year = map(int, selected_month.split('-'))
        month_name = datetime(year, month, 1).strftime('%B %Y')
    except ValueError:
        flash("Invalid month format. Using current month.", "error")
        current_date = datetime.now()
        month = current_date.month
        year = current_date.year
        month_name = current_date.strftime('%B %Y')

    # Get all working days in the selected month (Monday to Friday)
    working_days = calculate_working_days(year, month)
    working_days.sort()  # Ensure dates are in ascending order

    # Filter attendance for the selected month
    filtered_attendance = {}
    for name in all_names:
        filtered_attendance[name] = {}
    for record in attendance:
        name, date_str, in_time, out_time, status, expected_checkout, hours, day_status = record
        try:
            record_date = datetime.strptime(date_str, '%d/%m/%Y')
            if record_date.month == month and record_date.year == year:
                filtered_attendance[name][date_str] = record
        except ValueError:
            continue

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance - {month_name}"

    # Write headers
    # Row 1: Main headers (Name, dates, Total Days Present)
    main_headers = ['Name']
    for date in working_days:
        main_headers.extend([date, '', ''])
    main_headers.append(f"Total Days Present ({month_name})")
    ws.append(main_headers)

    # Row 2: Sub-headers (Overview, Check-in, Check-out under each date)
    sub_headers = ['']  # Empty cell under "Name"
    for _ in working_days:
        sub_headers.extend(['Overview', 'Check-in', 'Check-out'])
    sub_headers.append('')  # Empty cell under "Total Days Present"
    ws.append(sub_headers)

    # Merge cells for the date headers in Row 1
    for idx, date in enumerate(working_days, start=1):
        start_col = idx * 3 - 1  # Starting column for the date (e.g., B for first date)
        end_col = start_col + 2  # Merge 3 columns (e.g., B, C, D)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col)
        cell.alignment = Alignment(horizontal='center')

    # Apply formatting to headers
    for col_num in range(1, len(main_headers) + 1):
        cell_row1 = ws.cell(row=1, column=col_num)
        cell_row2 = ws.cell(row=2, column=col_num)
        cell_row1.font = Font(bold=True)
        cell_row2.font = Font(bold=True)
        cell_row1.alignment = Alignment(horizontal='center')
        cell_row2.alignment = Alignment(horizontal='center')

    # Populate data for each person
    for row_num, name in enumerate(all_names, start=3):
        row_data = [name]
        present_count = 0
        half_day_count = 0
        for date_str in working_days:
            record = filtered_attendance[name].get(date_str, None)
            if record:
                _, _, in_time, out_time, _, _, _, day_status = record
                in_time_str = in_time.split(':')[0] + ':' + in_time.split(':')[1] if in_time else '-'
                out_time_str = out_time.split(':')[0] + ':' + out_time.split(':')[1] if out_time else '-'
                if day_status == 'Full Day':
                    overview = 'p'
                    present_count += 1
                elif day_status == 'Half Day':
                    overview = 'h'
                    half_day_count += 1
                else:
                    overview = 'a'
                row_data.extend([overview, in_time_str, out_time_str])
            else:
                row_data.extend(['a', '-', '-'])  # Absent if no record

        # Calculate total days present: (p) + (h/2)
        total_days = present_count + (half_day_count / 2)
        row_data.append(f"{total_days:.1f}")
        ws.append(row_data)

    # Apply conditional formatting to the Overview columns
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    white_font = Font(color='FFFFFF')
    black_font = Font(color='000000')

    # Overview columns are at positions 2, 5, 8, ... (every 3rd column starting from B)
    for col_idx in range(2, len(working_days) * 3 + 2, 3):
        col_letter = get_column_letter(col_idx)
        # Apply formatting for Present (p) - green background, white text
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            FormulaRule(
                formula=[f'${col_letter}3="p"'],
                stopIfTrue=True,
                fill=green_fill,
                font=white_font
            )
        )
        # Apply formatting for Half Day (h) - red background, white text
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            FormulaRule(
                formula=[f'${col_letter}3="h"'],
                stopIfTrue=True,
                fill=red_fill,
                font=white_font
            )
        )
        # Apply formatting for Absent (a) - yellow background, black text
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            FormulaRule(
                formula=[f'${col_letter}3="a"'],
                stopIfTrue=True,
                fill=yellow_fill,
                font=black_font
            )
        )

    # Adjust column widths for better readability
    column_widths = {}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            # Skip MergedCell objects (fixed the module path here)
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(cell_value))
            except:
                pass
        column_widths[col_letter] = max_length + 2
        ws.column_dimensions[col_letter].width = column_widths[col_letter]

    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'attendance_{company_id}_{month_name.replace(" ", "_")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
@app.route('/history', methods=['GET', 'POST'])
def history():
    if session.get('role') != 'company_admin':
        return redirect(url_for('login'))

    company_id = session.get('company_id')
    selected_date = request.form.get('date') if request.method == 'POST' else None
    selected_user = request.form.get('name') if request.method == 'POST' else None
    selected_month = request.form.get('month') if request.method == 'POST' else None
    attendance_records = []
    user_history = []
    registered_users = list(load_encodings(company_id).keys())

    if not registered_users:
        flash("No registered users found.", "warning")
    else:
        flash(f"Found {len(registered_users)} registered users", "info")

    # Define working days (excluding weekends)
    def get_working_days(start_date, end_date):
        working_days = []
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:  # Monday to Friday
                working_days.append(current_date.strftime('%d/%m/%Y'))
            current_date += timedelta(days=1)
        return working_days

    # Fetch attendance data
    attendance = read_attendance_from_sheet(company_id)

    # Handle case when a specific date is selected
    if selected_date:
        try:
            formatted_date = datetime.strptime(selected_date, '%Y-%m-%d').strftime('%d/%m/%Y')
            # Check if the selected date is a working day
            if datetime.strptime(formatted_date, '%d/%m/%Y').weekday() >= 5:
                flash("Selected date is a weekend. No attendance records available.", "warning")
            else:
                # Generate records only for the selected date
                for user in registered_users:
                    user_records = [r for r in attendance if r[0] == user and r[1] == formatted_date]
                    if user_records:
                        record = user_records[0]
                        in_time = record[2] or ''
                        out_time = record[3] or ''
                        status = record[4] or 'Absent'
                        hours = record[6] or '0'
                        day_status = record[7] if len(record) > 7 else 'Absent'
                        time_display = f"{in_time} - {out_time}" if in_time and out_time else (in_time or out_time or 'N/A')
                        attendance_records.append({
                            'name': user,
                            'status': status,
                            'time': time_display,
                            'hours': hours,
                            'day_status': day_status
                        })
                    else:
                        # Add absent record for the user on the selected date
                        attendance_records.append({
                            'name': user,
                            'status': 'Absent',
                            'time': 'N/A',
                            'hours': '0',
                            'day_status': 'Absent'
                        })
        except ValueError:
            flash("Invalid date format.", "error")

    # Handle user history (only when no specific date is selected)
    elif selected_user:
        # Set date range (default: last 30 days or selected month)
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                start_date = datetime(year, month, 1)
                end_date = datetime(year, month, (datetime(year, month + 1, 1) - timedelta(days=1)).day)
            except ValueError:
                flash("Invalid month format.", "error")

        working_days = get_working_days(start_date, end_date)

        # Generate comprehensive attendance records for the date range
        comprehensive_attendance = []
        for user in registered_users:
            user_records = [r for r in attendance if r[0] == user]
            user_dates = set(r[1] for r in user_records)
            
            # Add existing records
            comprehensive_attendance.extend(user_records)
            
            # Add absent records for missing working days
            for date in working_days:
                if date not in user_dates:
                    comprehensive_attendance.append([
                        user,  # name
                        date,  # date
                        '',    # check_in
                        '',    # check_out
                        'Absent',  # status
                        '',    # (assuming unused)
                        '0',   # hours
                        'Absent'  # day_status
                    ])

        # Process user history for the selected user
        for record in comprehensive_attendance:
            if record[0] != selected_user:
                continue
            date = record[1]
            if selected_month:
                try:
                    record_month = datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m')
                    if record_month != selected_month:
                        continue
                except ValueError:
                    continue
            in_time = record[2] or ''
            out_time = record[3] or ''
            status = record[4] or 'Absent'
            hours = record[6] or '0'
            day_status = record[7] if len(record) > 7 else 'Absent'
            time_display = f"{in_time} - {out_time}" if in_time and out_time else (in_time or out_time or 'N/A')
            user_history.append({
                'date': date,
                'status': status,
                'time': time_display,
                'hours': hours,
                'day_status': day_status
            })

    return render_template('history.html', 
                           selected_date=selected_date,
                           selected_user=selected_user,
                           selected_month=selected_month,
                           attendance_records=attendance_records,
                           user_history=user_history,
                           registered_users=registered_users)


@app.route('/download_company_attendance')
def download_company_attendance():
    if session.get('role') != 'super_admin':
        return redirect(url_for('login'))

    # Get company_id and month from query parameters
    company_id = request.args.get('company_id')
    selected_month = request.args.get('month')

    if not company_id:
        flash("No company ID provided.", "error")
        return redirect(url_for('super_admin'))

    # Validate company_id exists
    result = safe_api_call(service.spreadsheets().values().get(spreadsheetId=SHEET_ID, range='Companies!A2:D'))
    companies = result.get('values', [])
    company = next((c for c in companies if c[0] == company_id), None)
    if not company:
        flash("Invalid company ID.", "error")
        return redirect(url_for('super_admin'))
    company_name = company[1]

    # Fetch attendance data for the company
    attendance = read_attendance_from_sheet(company_id)
    all_names = list(load_encodings(company_id).keys())

    # Default to current month if not specified
    if not selected_month:
        selected_month = datetime.now().strftime('%m-%Y')
    
    try:
        month, year = map(int, selected_month.split('-'))
        month_name = datetime(year, month, 1).strftime('%B %Y')
    except ValueError:
        flash("Invalid month format. Using current month.", "error")
        current_date = datetime.now()
        month = current_date.month
        year = current_date.year
        month_name = current_date.strftime('%B %Y')

    # Get all working days in the selected month (Monday to Friday)
    working_days = calculate_working_days(year, month)
    working_days.sort()  # Ensure dates are in ascending order

    # Filter attendance for the selected month
    filtered_attendance = {}
    for name in all_names:
        filtered_attendance[name] = {}
    for record in attendance:
        name, date_str, in_time, out_time, status, expected_checkout, hours, day_status = record
        try:
            record_date = datetime.strptime(date_str, '%d/%m/%Y')
            if record_date.month == month and record_date.year == year:
                filtered_attendance[name][date_str] = record
        except ValueError:
            continue

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance - {month_name}"

    # Write headers
    # Row 1: Main headers (Name, dates, Total Days Present)
    main_headers = ['Name']
    for date in working_days:
        main_headers.extend([date, '', ''])
    main_headers.append(f"Total Days Present ({month_name})")
    ws.append(main_headers)

    # Row 2: Sub-headers (Overview, Check-in, Check-out under each date)
    sub_headers = ['']  # Empty cell under "Name"
    for _ in working_days:
        sub_headers.extend(['Overview', 'Check-in', 'Check-out'])
    sub_headers.append('')  # Empty cell under "Total Days Present"
    ws.append(sub_headers)

    # Merge cells for the date headers in Row 1
    for idx, date in enumerate(working_days, start=1):
        start_col = idx * 3 - 1  # Starting column for the date (e.g., B for first date)
        end_col = start_col + 2  # Merge 3 columns (e.g., B, C, D)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col)
        cell.alignment = Alignment(horizontal='center')

    # Apply formatting to headers
    for col_num in range(1, len(main_headers) + 1):
        cell_row1 = ws.cell(row=1, column=col_num)
        cell_row2 = ws.cell(row=2, column=col_num)
        cell_row1.font = Font(bold=True)
        cell_row2.font = Font(bold=True)
        cell_row1.alignment = Alignment(horizontal='center')
        cell_row2.alignment = Alignment(horizontal='center')

    # Populate data for each person
    for row_num, name in enumerate(all_names, start=3):
        row_data = [name]
        present_count = 0
        half_day_count = 0
        for date_str in working_days:
            record = filtered_attendance[name].get(date_str, None)
            if record:
                _, _, in_time, out_time, _, _, _, day_status = record
                in_time_str = in_time.split(':')[0] + ':' + in_time.split(':')[1] if in_time else '-'
                out_time_str = out_time.split(':')[0] + ':' + out_time.split(':')[1] if out_time else '-'
                if day_status == 'Full Day':
                    overview = 'p'
                    present_count += 1
                elif day_status == 'Half Day':
                    overview = 'h'
                    half_day_count += 1
                else:
                    overview = 'a'
                row_data.extend([overview, in_time_str, out_time_str])
            else:
                row_data.extend(['a', '-', '-'])  # Absent if no record

        # Calculate total days present: (p) + (h/2)
        total_days = present_count + (half_day_count / 2)
        row_data.append(f"{total_days:.1f}")
        ws.append(row_data)

    # Apply conditional formatting to the Overview columns
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    white_font = Font(color='FFFFFF')
    black_font = Font(color='000000')

    # Overview columns are at positions 2, 5, 8, ... (every 3rd column starting from B)
    for col_idx in range(2, len(working_days) * 3 + 2, 3):
        col_letter = get_column_letter(col_idx)
        # Apply formatting for Present (p) - green background, white text
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            FormulaRule(
                formula=[f'${col_letter}3="p"'],
                stopIfTrue=True,
                fill=green_fill,
                font=white_font
            )
        )
        # Apply formatting for Half Day (h) - red background, white text
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            FormulaRule(
                formula=[f'${col_letter}3="h"'],
                stopIfTrue=True,
                fill=red_fill,
                font=white_font
            )
        )
        # Apply formatting for Absent (a) - yellow background, black text
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            FormulaRule(
                formula=[f'${col_letter}3="a"'],
                stopIfTrue=True,
                fill=yellow_fill,
                font=black_font
            )
        )

    # Adjust column widths for better readability
    column_widths = {}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(cell_value))
            except:
                pass
        column_widths[col_letter] = max_length + 2
        ws.column_dimensions[col_letter].width = column_widths[col_letter]

    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'attendance_{company_id}_{month_name.replace(" ", "_")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)